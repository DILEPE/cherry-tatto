"""Exportación XLSX del informe financiero de citas (panel Streamlit)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Optional

from app.domain.appointment_money import (
    appointment_financial_totals,
    customer_credit_from_row,
)
from streamlit_app.appointment_staff_labels import assigned_artist_display_name


def _xlsx_border_thin() -> Any:
    from openpyxl.styles import Border, Side

    s = Side(style="thin", color="FF9CA3AF")
    return Border(left=s, right=s, top=s, bottom=s)


def _excel_set_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    *,
    font=None,
    alignment=None,
    border=None,
    fill=None,
) -> None:
    """Asignación de celda compatible con coordenadas 1-based tipo Excel."""
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill


def _excel_apply_border_block(
    ws: Any, row_min: int, row_max: int, col_min: int, col_max: int, border: Any
) -> None:
    for r in range(row_min, row_max + 1):
        for c in range(col_min, col_max + 1):
            ws.cell(row=r, column=c).border = border


def citas_filtered_to_excel_bytes(
    rows: list[dict[str, Any]],
    *,
    generated_at: Optional[datetime] = None,
) -> bytes:
    """Genera .xlsx financiero estilizado: título, encabezados en negrita y tablas demarcadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    gen_dt = generated_at or datetime.now()
    fecha_etiqueta = gen_dt.strftime("%d/%m/%Y %H:%M")

    datos: list[list[Any]] = []
    for r in rows:
        tot, abo, pend = appointment_financial_totals(r)
        cred = customer_credit_from_row(r)
        nombre = str(r.get("customer_name") or r.get("name") or "").strip()
        artista = assigned_artist_display_name(r)
        datos.append(
            [
                nombre,
                artista,
                round(tot, 2),
                round(abo, 2),
                round(pend, 2),
                round(cred, 2),
            ]
        )

    headers = [
        "Cliente",
        "Artista / profesional",
        "Valor total (COP)",
        "Abonado (COP)",
        "Pendiente (COP)",
        "Saldo a favor (COP)",
    ]
    ncol = len(headers)

    wb = Workbook()
    bd = _xlsx_border_thin()
    font_title = Font(bold=True, size=14, color="FF111827")
    font_sub = Font(size=10, color="FF4B5563")
    font_header = Font(bold=True, size=11, color="FF111827")
    font_body = Font(size=10, color="FF374151")
    fill_header = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right_num = Alignment(horizontal="right", vertical="center")

    ws1 = wb.active
    ws1.title = "Datos financieros"
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws1.cell(1, 1, "Informe financiero — Citas Cherry Ink · Rock City")
    ws1.cell(1, 1).font = font_title
    ws1.cell(1, 1).alignment = align_center

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws1.cell(2, 1, f"Generado: {fecha_etiqueta}")
    ws1.cell(2, 1).font = font_sub
    ws1.cell(2, 1).alignment = align_center

    header_row = 4
    for col in range(1, ncol + 1):
        _excel_set_cell(
            ws1,
            header_row,
            col,
            headers[col - 1],
            font=font_header,
            fill=fill_header,
            alignment=align_left if col <= 2 else align_right_num,
            border=bd,
        )

    row_start_body = header_row + 1
    if datos:
        for i, row_vals in enumerate(datos):
            r = row_start_body + i
            for col in range(1, ncol + 1):
                v = row_vals[col - 1]
                align = align_left if col <= 2 else align_right_num
                _excel_set_cell(ws1, r, col, v, font=font_body, alignment=align, border=bd)
            rmax = r
        _excel_apply_border_block(ws1, header_row, rmax, 1, ncol, bd)
    else:
        ws1.merge_cells(
            start_row=row_start_body, start_column=1, end_row=row_start_body, end_column=ncol
        )
        c_msg = ws1.cell(row=row_start_body, column=1, value="Sin filas para los filtros actuales.")
        c_msg.font = font_body
        c_msg.alignment = align_left
        _excel_apply_border_block(ws1, header_row, row_start_body, 1, ncol, bd)

    for col in range(1, ncol + 1):
        letter = get_column_letter(col)
        w = 18
        if col == 1:
            w = 34
        elif col == 2:
            w = 28
        ws1.column_dimensions[letter].width = w

    ws2 = wb.create_sheet("Resumen financiero", 1)
    rtot = rabo = rpend = rfav = 0.0
    for rr in rows:
        t, a, p = appointment_financial_totals(rr)
        rtot += t
        rabo += a
        rpend += p
        rfav += customer_credit_from_row(rr)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws2.cell(1, 1, "Resumen financiero — mismos filtros que el panel")
    ws2.cell(1, 1).font = font_title
    ws2.cell(1, 1).alignment = align_center
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws2.cell(2, 1, f"Generado: {fecha_etiqueta}")
    ws2.cell(2, 1).font = font_sub
    ws2.cell(2, 1).alignment = align_center

    resumen_labels = [
        "Total valor trabajo (COP)",
        "Total abonado (COP)",
        "Total pendiente (COP)",
        "Total saldo a favor (COP)",
        "Cantidad de citas",
    ]
    resumen_vals: list[Any] = [
        round(rtot, 2),
        round(rabo, 2),
        round(rpend, 2),
        round(rfav, 2),
        len(rows) if rows else 0,
    ]
    hdr_r = 4
    _excel_set_cell(
        ws2, hdr_r, 1, "Concepto", font=font_header, fill=fill_header, alignment=align_left, border=bd
    )
    _excel_set_cell(
        ws2, hdr_r, 2, "Valor", font=font_header, fill=fill_header, alignment=align_right_num, border=bd
    )
    for i, lab in enumerate(resumen_labels):
        r = hdr_r + 1 + i
        _excel_set_cell(ws2, r, 1, lab, font=font_body, alignment=align_left, border=bd)
        _excel_set_cell(ws2, r, 2, resumen_vals[i], font=font_body, alignment=align_right_num, border=bd)
    _excel_apply_border_block(ws2, hdr_r, hdr_r + len(resumen_labels), 1, 2, bd)
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


__all__ = ["citas_filtered_to_excel_bytes"]
