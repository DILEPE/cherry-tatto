"""Exportación XLSX del listado de clientes (panel Streamlit)."""

from __future__ import annotations

import hashlib
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from app.schemas.customer import CUSTOMER_BIRTH_PENDING
from streamlit_app import api_client


def _xlsx_border_thin() -> Any:
    from openpyxl.styles import Border, Side

    s = Side(style="thin", color="FF9CA3AF")
    return Border(left=s, right=s, top=s, bottom=s)


def _excel_set_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    *,
    font=None,
    alignment=None,
    border=None,
    fill=None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill


def _excel_apply_border_block(
    ws: Any, row_min: int, row_max: int, col_min: int, col_max: int, border: Any
) -> None:
    for r in range(row_min, row_max + 1):
        for c in range(col_min, col_max + 1):
            ws.cell(row=r, column=c).border = border


def _fmt_date(val: Any) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        d = val.date()
    elif isinstance(val, date):
        d = val
    else:
        s = str(val).strip()[:10]
        if not s:
            return ""
        try:
            d = datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return s
    if d == CUSTOMER_BIRTH_PENDING:
        return "Pendiente"
    return d.strftime("%d/%m/%Y")


def _fmt_datetime(val: Any) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    s = str(val).strip().replace("T", " ")
    if not s:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s[:19] if len(s) > 10 else s[:10], fmt)
            if fmt == "%Y-%m-%d":
                return dt.strftime("%d/%m/%Y")
            return dt.strftime("%d/%m/%Y %H:%M")
        except ValueError:
            pass
    return s[:16]


def _yes_no_minor(val: Any) -> str:
    if val in (True, 1, "1", "true", "True"):
        return "Sí"
    if val in (False, 0, "0", "false", "False"):
        return "No"
    return "—"


def fetch_all_customers_for_export(
    search: str = "",
    *,
    batch_size: int = 100,
) -> tuple[list[dict[str, Any]], int, Optional[str]]:
    """Pagina GET /api/customers (máx. 100 filas por petición, límite de la API)."""
    """Descarga todas las filas que coinciden con el filtro de búsqueda actual."""
    term = (search or "").strip()
    offset = 0
    collected: list[dict[str, Any]] = []
    total = 0
    while True:
        params_search = term or None
        ok, code, data = api_client.get_customers(
            limit=batch_size,
            offset=offset,
            search=params_search,
        )
        if not ok or not isinstance(data, dict):
            err = data
            if isinstance(err, dict):
                err = err.get("detail", err)
            return [], 0, f"HTTP {code}: {err}"
        total = int(data.get("total") or 0)
        chunk = [x for x in (data.get("items") or []) if isinstance(x, dict)]
        collected.extend(chunk)
        if len(collected) >= total or len(chunk) < batch_size:
            break
        offset += batch_size
    return collected, total, None


def customer_row_to_excel_values(c: dict[str, Any]) -> list[Any]:
    return [
        int(c.get("id") or 0),
        str(c.get("first_name") or "").strip(),
        str(c.get("last_name") or "").strip(),
        _fmt_date(c.get("birth_date")),
        str(c.get("document_type") or "").strip(),
        str(c.get("document_number") or "").strip(),
        _fmt_date(c.get("document_issue_date")),
        str(c.get("email") or "").strip(),
        str(c.get("phone_number") or "").strip(),
        str(c.get("address") or "").strip(),
        str(c.get("nationality") or "").strip(),
        str(c.get("profession") or "").strip(),
        str(c.get("social_media") or "").strip(),
        str(c.get("emergency_contact_name") or "").strip(),
        str(c.get("emergency_contact_phone") or "").strip(),
        _yes_no_minor(c.get("is_minor")),
        str(c.get("guardian_name") or "").strip(),
        str(c.get("guardian_document_type") or "").strip(),
        str(c.get("guardian_document_number") or "").strip(),
        _fmt_date(c.get("guardian_document_issue_date")),
        _fmt_datetime(c.get("created_at")),
        _fmt_datetime(c.get("updated_at")),
    ]


CUSTOMERS_EXCEL_HEADERS: tuple[str, ...] = (
    "ID",
    "Nombres",
    "Apellidos",
    "Fecha nacimiento",
    "Tipo documento",
    "Número documento",
    "Fecha expedición documento",
    "Correo",
    "Teléfono",
    "Dirección",
    "Nacionalidad",
    "Profesión",
    "Redes sociales",
    "Contacto emergencia (nombre)",
    "Contacto emergencia (teléfono)",
    "Menor de edad",
    "Tutor (nombre)",
    "Tutor (tipo documento)",
    "Tutor (número documento)",
    "Tutor (fecha expedición)",
    "Fecha registro",
    "Última actualización",
)


def customers_to_excel_bytes(
    rows: list[dict[str, Any]],
    *,
    search_label: str = "",
    generated_at: Optional[datetime] = None,
) -> bytes:
    """Genera .xlsx con datos de clientes y hoja resumen."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    gen_dt = generated_at or datetime.now()
    fecha_etiqueta = gen_dt.strftime("%d/%m/%Y %H:%M")
    filtro = (search_label or "").strip() or "Todos"

    datos = [customer_row_to_excel_values(c) for c in rows]
    headers = list(CUSTOMERS_EXCEL_HEADERS)
    ncol = len(headers)

    wb = Workbook()
    bd = _xlsx_border_thin()
    font_title = Font(bold=True, size=14, color="FF111827")
    font_sub = Font(size=10, color="FF4B5563")
    font_header = Font(bold=True, size=11, color="FF111827")
    font_body = Font(size=10, color="FF374151")
    fill_header = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Clientes"
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws1.cell(1, 1, "Informe de clientes — Cherry Ink · Rock City")
    ws1.cell(1, 1).font = font_title
    ws1.cell(1, 1).alignment = align_center

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws1.cell(2, 1, f"Generado: {fecha_etiqueta} · Filtro búsqueda: {filtro}")
    ws1.cell(2, 1).font = font_sub
    ws1.cell(2, 1).alignment = align_center

    header_row = 4
    for col in range(1, ncol + 1):
        _excel_set_cell(
            ws1,
            header_row,
            col,
            headers[col - 1],
            font=font_header,
            fill=fill_header,
            alignment=align_left,
            border=bd,
        )

    row_start_body = header_row + 1
    if datos:
        for i, row_vals in enumerate(datos):
            r = row_start_body + i
            for col in range(1, ncol + 1):
                _excel_set_cell(ws1, r, col, row_vals[col - 1], font=font_body, alignment=align_left, border=bd)
            rmax = r
        _excel_apply_border_block(ws1, header_row, rmax, 1, ncol, bd)
    else:
        ws1.merge_cells(
            start_row=row_start_body, start_column=1, end_row=row_start_body, end_column=ncol
        )
        c_msg = ws1.cell(row=row_start_body, column=1, value="Sin clientes para el filtro actual.")
        c_msg.font = font_body
        c_msg.alignment = align_left
        _excel_apply_border_block(ws1, header_row, row_start_body, 1, ncol, bd)

    col_widths = {
        1: 8,
        2: 18,
        3: 18,
        4: 14,
        5: 12,
        6: 16,
        7: 16,
        8: 28,
        9: 14,
        10: 32,
        11: 14,
        12: 18,
        13: 22,
        14: 22,
        15: 18,
        16: 10,
        17: 22,
        18: 14,
        19: 16,
        20: 16,
        21: 18,
        22: 18,
    }
    for col in range(1, ncol + 1):
        ws1.column_dimensions[get_column_letter(col)].width = col_widths.get(col, 16)

    ws2 = wb.create_sheet("Resumen", 1)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws2.cell(1, 1, "Resumen — clientes exportados")
    ws2.cell(1, 1).font = font_title
    ws2.cell(1, 1).alignment = align_center
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws2.cell(2, 1, f"Generado: {fecha_etiqueta}")
    ws2.cell(2, 1).font = font_sub
    ws2.cell(2, 1).alignment = align_center

    menores = sum(1 for c in rows if c.get("is_minor") in (True, 1, "1"))
    resumen_labels = [
        "Total clientes en exportación",
        "Menores de edad",
        "Filtro de búsqueda aplicado",
    ]
    resumen_vals: list[Any] = [len(rows), menores, filtro]
    hdr_r = 4
    align_right = Alignment(horizontal="right", vertical="center")
    _excel_set_cell(ws2, hdr_r, 1, "Concepto", font=font_header, fill=fill_header, alignment=align_left, border=bd)
    _excel_set_cell(ws2, hdr_r, 2, "Valor", font=font_header, fill=fill_header, alignment=align_right, border=bd)
    for i, lab in enumerate(resumen_labels):
        r = hdr_r + 1 + i
        _excel_set_cell(ws2, r, 1, lab, font=font_body, alignment=align_left, border=bd)
        _excel_set_cell(ws2, r, 2, resumen_vals[i], font=font_body, alignment=align_right, border=bd)
    _excel_apply_border_block(ws2, hdr_r, hdr_r + len(resumen_labels), 1, 2, bd)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 36

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def customers_excel_cache_fingerprint(search: str, total: int) -> str:
    norm = (search or "").strip().lower()
    return hashlib.md5(f"{norm}|{int(total)}".encode()).hexdigest()[:16]


def build_customers_excel_for_export(search: str) -> tuple[bytes, Optional[str]]:
    """Genera bytes del informe; devuelve error si falla la API."""
    rows, _total, err = fetch_all_customers_for_export(search)
    if err:
        return b"", err
    try:
        data = customers_to_excel_bytes(rows, search_label=search, generated_at=datetime.now())
    except Exception as e:
        return b"", str(e)
    return data, None


__all__ = [
    "CUSTOMERS_EXCEL_HEADERS",
    "build_customers_excel_for_export",
    "customers_excel_cache_fingerprint",
    "customers_to_excel_bytes",
    "fetch_all_customers_for_export",
]
