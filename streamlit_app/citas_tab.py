"""Streamlit: Citas con calendario, franjas horarias y formulario mínimo."""
from __future__ import annotations

import calendar
import hashlib
import html as html_mod
import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from io import BytesIO
from collections import defaultdict
from typing import Any, Dict, List, Optional

import streamlit as st
import streamlit.components.v1 as components

from pydantic import ValidationError

from app.domain.service_types import resolve_service_type
from app.domain.contract_kinds import (
    SCOPE_LABEL_ES,
    appointment_to_contract_kind,
    service_type_requires_contract,
    service_type_to_contract_kind,
)
from app.domain.contract_signing_guard import appointment_must_be_fully_paid_for_contract
from app.domain.survey_question_helpers import question_type_label_es, question_type_supports_distribution_chart
from app.schemas.customer import CUSTOMER_BIRTH_PENDING, CustomerCreate
from streamlit_app import api_client, report_charts
from streamlit_app.panel_navigation import (
    open_contract_artist_signature,
    open_contract_express_piercing,
    open_contract_signing,
)
from streamlit_app.cached_public_api import (
    get_panel_users_assignable_cached,
    get_survey_question_stats_summary_cached,
)
from streamlit_app.customer_sync import fetch_customer_by_document
from streamlit_app.validation import validate_appointment

from streamlit_app.components.citas_legend import render_citas_color_legend as _render_citas_color_legend
from streamlit_app.components.pills import (
    client_history_key as _client_history_key,
    client_pill_class as _client_pill_class,
    customer_name_pill_html as _customer_name_pill_html,
    row_is_priority as _row_is_priority,
    status_pill_html as _status_pill_html,
)
from streamlit_app.components.service_flags import service_type_flag_html as _service_type_flag_html
from streamlit_app.styles.inject import inject_via_streamlit_lazy


def _booking_customer_create_for_existing_client(
    snap: Dict[str, Any],
    *,
    first_name: str,
    last_name: str,
    phone_number: str,
    email_s: str,
    document_number: str,
) -> CustomerCreate:
    """Fusiona la ficha API (`snap`) con nombre/teléfono/correo editados en el formulario de agendamiento."""
    bd_raw = snap.get("birth_date")
    if isinstance(bd_raw, str) and bd_raw.strip():
        birth_date = date.fromisoformat(bd_raw.strip()[:10])
    elif isinstance(bd_raw, date):
        birth_date = bd_raw
    elif isinstance(bd_raw, datetime):
        birth_date = bd_raw.date()
    else:
        birth_date = CUSTOMER_BIRTH_PENDING

    doc_issue_d: Optional[date] = None
    doc_issue = snap.get("document_issue_date")
    if doc_issue is not None and str(doc_issue).strip():
        if isinstance(doc_issue, str):
            doc_issue_d = date.fromisoformat(str(doc_issue).strip()[:10])
        elif isinstance(doc_issue, date):
            doc_issue_d = doc_issue
        elif isinstance(doc_issue, datetime):
            doc_issue_d = doc_issue.date()

    raw_ty = str(snap.get("document_type") or "CC").strip().upper()
    if raw_ty not in ("CC", "TI", "CE", "PAS"):
        raw_ty = "CC"

    g_issue: Optional[date] = None
    g_raw = snap.get("guardian_document_issue_date")
    if g_raw is not None and str(g_raw).strip():
        if isinstance(g_raw, str):
            g_issue = date.fromisoformat(str(g_raw).strip()[:10])
        elif isinstance(g_raw, date):
            g_issue = g_raw
        elif isinstance(g_raw, datetime):
            g_issue = g_raw.date()

    gdt_clean: Optional[str] = None
    gdt = snap.get("guardian_document_type")
    if gdt is not None and str(gdt).strip():
        u = str(gdt).strip().upper()
        if u in ("CC", "TI", "CE", "PAS"):
            gdt_clean = u

    sm_raw = snap.get("social_media")
    social_media: Optional[str] = None
    if isinstance(sm_raw, str) and sm_raw.strip():
        social_media = sm_raw.strip()
    elif sm_raw is not None and not isinstance(sm_raw, str):
        social_media = str(sm_raw).strip() or None

    return CustomerCreate(
        first_name=first_name.strip(),
        last_name=last_name.strip(),
        birth_date=birth_date,
        document_type=raw_ty,  # type: ignore[arg-type]
        document_number=document_number.strip(),
        document_issue_date=doc_issue_d,
        email=email_s,
        phone_number=phone_number.strip(),
        address=(str(snap["address"]).strip() if snap.get("address") else None),
        nationality=(str(snap["nationality"]).strip() if snap.get("nationality") else None),
        profession=(str(snap["profession"]).strip() if snap.get("profession") else None),
        social_media=social_media,
        emergency_contact_name=(
            str(snap["emergency_contact_name"]).strip() if snap.get("emergency_contact_name") else None
        ),
        emergency_contact_phone=(
            str(snap["emergency_contact_phone"]).strip() if snap.get("emergency_contact_phone") else None
        ),
        is_minor=bool(snap.get("is_minor")),
        guardian_name=(str(snap["guardian_name"]).strip() if snap.get("guardian_name") else None),
        guardian_document_type=gdt_clean,  # type: ignore[arg-type]
        guardian_document_number=(
            str(snap["guardian_document_number"]).strip() if snap.get("guardian_document_number") else None
        ),
        guardian_document_issue_date=g_issue,
    )


def _api_error(payload: Any) -> str:
    if isinstance(payload, dict):
        return str(payload.get("detail", payload))
    return str(payload)


_AP_ACTION_INFO_KEY = "_ap_action_info"
_AP_TOAST_FETCH_ERR_KEY = "_ap_toast_fetch_err"
_AP_TOAST_FIN_SAVE_ERR_KEY = "_ap_toast_fin_save_err"


def _queue_appointment_action_success(msg: str) -> None:
    """Confirmación visible en la siguiente ejecución (pestaña Citas o Reporte)."""
    st.session_state[_AP_ACTION_INFO_KEY] = msg


def _render_appointment_action_feedback() -> None:
    msg = st.session_state.pop(_AP_ACTION_INFO_KEY, None)
    if msg:
        st.toast(msg, icon="✅", duration="long")


def _render_appointments_fetch_error_toast() -> None:
    """Un toast por mensaje de fallo al cargar citas (evita repetir en cada rerun)."""
    err = st.session_state.get("_ap_err")
    if not err:
        st.session_state.pop(_AP_TOAST_FETCH_ERR_KEY, None)
        return
    if st.session_state.get(_AP_TOAST_FETCH_ERR_KEY) != err:
        st.session_state[_AP_TOAST_FETCH_ERR_KEY] = err
        st.toast(str(err), icon="❌", duration="long")


def _toast_financial_save_error_if_any() -> None:
    save_err = st.session_state.get("_ap_fin_save_error")
    if not save_err:
        st.session_state.pop(_AP_TOAST_FIN_SAVE_ERR_KEY, None)
        return
    if st.session_state.get(_AP_TOAST_FIN_SAVE_ERR_KEY) != save_err:
        st.session_state[_AP_TOAST_FIN_SAVE_ERR_KEY] = save_err
        st.toast(str(save_err), icon="❌", duration="long")


def _format_dt_for_user_message(dt_str: str) -> str:
    """Presentación corta para mensajes (YYYY-MM-DD HH:MM a DD/MM/YYYY HH:MM)."""
    raw = (dt_str or "").strip().replace("T", " ")[:16]
    if len(raw) < 16:
        return raw or "—"
    try:
        dt = datetime.strptime(raw, "%Y-%m-%d %H:%M")
        return dt.strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return raw


def _may_see_all_appointments() -> bool:
    """Vendedor / administrador / modo env con acceso total ven el listado completo (con filtro opcional)."""
    from streamlit_app.panel_auth import panel_auth_enabled

    if not panel_auth_enabled():
        return True
    if st.session_state.get("_panel_session_full_access"):
        return True
    role = str(st.session_state.get("_panel_user_role") or "")
    return role in ("administrador", "vendedor")


def _panel_is_technician_role() -> bool:
    """Tatuador o perforador: citas activas propias desde hoy; no agenda ni montos ni reprogramar."""
    role = str(st.session_state.get("_panel_user_role") or "")
    return role in ("tatuador", "perforador")


def _appointment_row_active_for_technician(row: dict[str, Any]) -> bool:
    stv = str(row.get("status") or "").strip().lower()
    return stv in ("agendada", "reprogramada")


def _appointment_for_technician_visible_date(row: dict[str, Any], *, ref_day: date) -> bool:
    """Solo citas con fecha de agenda >= día de referencia (típicamente hoy)."""
    ad = _parse_date(row.get("appointment_date", row.get("date")))
    return ad >= ref_day


def _filter_appointments_for_session_role(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Técnicos: estados activos y fecha de cita desde hoy en adelante."""
    if not _panel_is_technician_role():
        return items
    today = date.today()
    return [
        r
        for r in items
        if _appointment_row_active_for_technician(r)
        and _appointment_for_technician_visible_date(r, ref_day=today)
    ]


def _technician_clear_disallowed_dialog_states() -> None:
    """Evita abrir agendar / montos / reprogramar / recibos / anular si el rol es solo técnico."""
    if not _panel_is_technician_role():
        return
    st.session_state.pop("_ap_dlg", None)
    st.session_state.pop("_ap_fin_item", None)
    st.session_state.pop("_ap_reprogram_item", None)
    st.session_state.pop("_ap_cancel_item", None)
    st.session_state.pop("_ap_receipts_item", None)


def _clear_calendar_dialog_focus() -> None:
    """Cierra diálogos de calendario asociados al día o a una cita concreta."""
    st.session_state.pop("_cal_overflow_day", None)
    st.session_state.pop("_cal_focus_appt_id", None)


def _find_appointment_row_by_id(appt_id: int) -> Optional[dict[str, Any]]:
    for row in st.session_state.get("_ap_list") or []:
        if int(row.get("id", 0) or 0) == int(appt_id):
            return row
    return None


def _contract_firma_blocked_por_saldo(r: Dict[str, Any]) -> bool:
    """True si hay valor total en la cita y aún queda saldo pendiente (no se puede firmar)."""
    ok, _ = appointment_must_be_fully_paid_for_contract(
        total_amount=r.get("total_amount"),
        deposit=r.get("deposit"),
        pending_balance=r.get("pending_balance"),
    )
    return not ok


def _firmar_contrato_disabled(r: Dict[str, Any]) -> bool:
    """Recepción no reabre tras registrar contrato; el técnico sí puede completar su firma pendiente."""
    appt_id = int(r.get("id", 0) or 0)
    status = str(r.get("status") or "Agendada")
    has_customer = r.get("customer_id") is not None
    base = (
        appt_id <= 0
        or not has_customer
        or status in {"Cancelada", "Finalizada"}
        or not service_type_requires_contract(str(r.get("service_type") or ""))
        or _contract_firma_blocked_por_saldo(r)
    )
    if base:
        return True
    has_contract = bool(r.get("has_signed_contract"))
    pending_artist = bool(r.get("contract_pending_artist_signature"))
    if _panel_is_technician_role():
        # El técnico solo completa su firma sobre un contrato ya guardado en recepción (sin datos ni encuesta).
        return base or not pending_artist
    return has_contract


def _firmar_contrato_button_label(r: Dict[str, Any]) -> str:
    """Texto del botón según estado del contrato (requiere `contract_pending_artist_signature` en la API)."""
    has_contract = bool(r.get("has_signed_contract"))
    pending_artist = bool(r.get("contract_pending_artist_signature"))
    if _panel_is_technician_role():
        if pending_artist:
            return "Completar firma profesional"
        if has_contract:
            return "Contrato firmado"
        return "Firma pendiente en recepción"
    if not has_contract:
        return "Firmar contrato"
    if pending_artist:
        return "Pendiente firma profesional"
    return "Contrato firmado"


def _open_firma_contrato_nav(r: Dict[str, Any], appt_id: int) -> None:
    if _panel_is_technician_role():
        if bool(r.get("contract_pending_artist_signature")):
            open_contract_artist_signature(appt_id)
        return
    open_contract_signing(appt_id)


def _appointments_query_assigned_user_id() -> Optional[int]:
    if not _may_see_all_appointments():
        uid = st.session_state.get("_panel_user_id")
        return int(uid) if uid is not None else None
    raw = st.session_state.get("_ap_filter_artist_id")
    if raw is None or raw == 0:
        return None
    return int(raw)


def _ensure_assignable_staff() -> list[dict[str, Any]]:
    cached = st.session_state.get("_ap_assignable_staff")
    if isinstance(cached, list):
        return cached
    ok, _, data = get_panel_users_assignable_cached()
    if ok and isinstance(data, list):
        st.session_state["_ap_assignable_staff"] = data
        return data
    return []


def _work_kind_to_assignee_role(work_kind: str) -> str:
    if work_kind == "tatuaje":
        return "tatuador"
    return "perforador"


def _work_kind_to_schedule_kind(work_kind: str) -> str:
    """
    Eje de agenda: solo sesión de tatuaje vs todo lo de piercing (colocación, limpieza, cambio).
    Las franjas de un eje no bloquean al otro.
    """
    if work_kind == "tatuaje":
        return "tattoo"
    return "piercing"


def _appointments_for_artist_schedule(
    items: list[dict[str, Any]],
    day: date,
    artist_id: Optional[int],
    *,
    schedule_kind: str,
    exclude_appointment_id: Optional[int] = None,
) -> list[dict[str, Any]]:
    """
    Citas que compiten por huecos: mismo profesional (o sin asignar en legacy, todas las ramas)
    y mismo tipo de agenda (`tattoo` vs `piercing`).
    """
    out: list[dict[str, Any]] = []
    for row in _appointments_same_day_raw(items, day):
        rid = int(row.get("id") or 0)
        if exclude_appointment_id is not None and rid == exclude_appointment_id:
            continue
        if str(row.get("status") or "").strip().lower() == "cancelada":
            continue
        if appointment_to_contract_kind(row) != schedule_kind:
            continue
        ra = row.get("assigned_panel_user_id")
        if ra is None or ra == "":
            out.append(row)
        elif artist_id is not None and int(ra) == int(artist_id):
            out.append(row)
    return out


def _appointments_same_day_schedule_kind(
    items: list[dict[str, Any]],
    day: date,
    schedule_kind: str,
) -> list[dict[str, Any]]:
    """Mismo día y eje tatuaje/piercing (sin filtrar por profesional; p. ej. falta asignación)."""
    out: list[dict[str, Any]] = []
    for row in _appointments_same_day_raw(items, day):
        if str(row.get("status") or "").strip().lower() == "cancelada":
            continue
        if appointment_to_contract_kind(row) != schedule_kind:
            continue
        out.append(row)
    return out


def _assigned_staff_label(row: dict[str, Any]) -> str:
    fn = str(row.get("assigned_first_name") or "").strip()
    ln = str(row.get("assigned_last_name") or "").strip()
    un = str(row.get("assigned_username") or "").strip()
    if not fn and not ln and not un:
        return "—"
    name = f"{fn} {ln}".strip()
    if name and un:
        return f"{name} (@{un})"
    if name:
        return name
    return f"@{un}" if un else "—"


def _assigned_artist_display_name(row: dict[str, Any]) -> str:
    """Nombre del artista/profesional asignado (nombre y apellido; si no hay, usuario de panel)."""
    fn = str(row.get("assigned_first_name") or "").strip()
    ln = str(row.get("assigned_last_name") or "").strip()
    name = f"{fn} {ln}".strip()
    if name:
        return name
    un = str(row.get("assigned_username") or "").strip()
    if un:
        return f"@{un}"
    return "Sin asignar"


def _artist_filter_labels_and_map() -> tuple[list[str], dict[str, int]]:
    from app.domain.panel_user_profile import PANEL_ROLE_LABEL_ES

    staff = _ensure_assignable_staff()
    labels: list[str] = ["Todos"]
    id_by_label: dict[str, int] = {"Todos": 0}
    for s in staff:
        r = str(s.get("role") or "")
        tag = PANEL_ROLE_LABEL_ES.get(r, r)
        lab = (
            f"{s.get('first_name', '')} {s.get('last_name', '')} (@{s.get('username', '')}) — {tag}"
        ).strip()
        if lab in id_by_label:
            lab = f"{lab} · id {s.get('id')}"
        labels.append(lab)
        id_by_label[lab] = int(s["id"])
    return labels, id_by_label


def _render_professional_calendar_filter() -> None:
    """Filtro por tatuador/perforador para quien puede ver toda la agenda."""
    from app.domain.panel_user_profile import PANEL_ROLE_LABEL_ES

    if not _may_see_all_appointments():
        if _panel_is_technician_role():
            st.caption(
                "Solo ves citas **activas** (agendada / reprogramada) **asignadas a ti**, "
                "**desde la fecha de hoy en adelante** (las pasadas no aparecen). "
                f"Rol: **{PANEL_ROLE_LABEL_ES.get(str(st.session_state.get('_panel_user_role') or ''), 'operador')}**."
            )
        else:
            st.caption(
                "Solo ves citas asignadas a **tu usuario** del panel ("
                f"{PANEL_ROLE_LABEL_ES.get(str(st.session_state.get('_panel_user_role') or ''), 'operador')})."
            )
        st.session_state["_ap_filter_artist_id"] = 0
        return
    labels, id_by_label = _artist_filter_labels_and_map()
    sb_key = "_ap_filt_artist_cal"
    if sb_key not in st.session_state:
        st.session_state[sb_key] = "Todos"
    choice = st.selectbox(
        "Profesional (filtro de citas)",
        options=labels,
        key=sb_key,
        help="Filtra qué citas se cargan desde la API. "
        "Administrador y vendedor ven la agenda completa con este filtro; "
        "tatuadores y perforadores solo cargan sus citas activas asignadas.",
    )
    st.session_state["_ap_filter_artist_id"] = id_by_label.get(str(choice), 0)


def _reprogram_disabled_for_row(r: Dict[str, Any]) -> bool:
    """Reprogramar solo en Agendada/Reprogramada, sin contrato firmado y no cancelada."""
    appt_id = int(r.get("id", 0) or 0)
    status = str(r.get("status") or "Agendada")
    if appt_id <= 0 or status == "Cancelada":
        return True
    if status not in {"Agendada", "Reprogramada"}:
        return True
    if bool(r.get("has_signed_contract")):
        return True
    return False


def _show_validation_errors(errors: List[Any]) -> None:
    for e in errors:
        st.markdown(
            f'<div class="m-error"><strong>{e.field}</strong>: {e.message}</div>',
            unsafe_allow_html=True,
        )


def _format_cop(value: float | int) -> str:
    amount = int(round(float(value or 0)))
    return f"COP ${amount:,.0f}".replace(",", ".")


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _financial_row_values(row: dict[str, Any]) -> tuple[float, float, float]:
    """
    Normaliza montos para UI y resumen:
    - total nunca menor que abonado (fallback datos legacy)
    - pendiente: si viene `pending_balance` de la API/MySQL es la fuente de verdad
      (ej. tras anular con saldo ya puesto en 0 y crédito en otra columna);
      si no, pendiente = max(total − abonado − saldo a favor, 0) para no ignorar créditos.
    """
    abonado = max(_to_float(row.get("deposit"), 0.0), 0.0)
    total_raw = max(_to_float(row.get("total_amount"), 0.0), 0.0)
    total = max(total_raw, abonado)
    cred = max(_to_float(row.get("customer_credit"), 0.0), 0.0)
    raw_pb = row.get("pending_balance")
    if raw_pb is not None and raw_pb != "":
        pendiente = max(round(_to_float(raw_pb, 0.0), 2), 0.0)
    else:
        pendiente = max(round(total - abonado - cred, 2), 0.0)
    return total, abonado, pendiente


def _calendar_month_value_label(total: float) -> str:
    """
    Total en celda del calendario mensual: solo dígitos (sin símbolo $ ni sufijo tipo «k»).
    Por debajo de mil se muestra el valor tal cual (unidades hasta 999).
    A partir de 1.000 se trunca quitando los tres últimos dígitos (− unidades/decenas/centenas),
    de modo que el número mostrado queda «en miles de pesos» sin escribir esos tres ceros.
    Si ese número tiene 4+ dígitos, se agrupa con punto como miles (p. ej. 5_678).
    Para montos ≥ 1.000.000 el mismo //1000 deja también fuera los tres dígitos de menores orden.
    """
    v = int(round(max(float(total or 0), 0)))
    if v <= 0:
        return "—"
    if v < 1000:
        return str(v)
    n = v // 1000
    if n < 1000:
        return str(n)
    return f"{n:,.0f}".replace(",", ".")


def _customer_credit_value(row: dict[str, Any]) -> float:
    """Saldo a favor del cliente asociado a esta cita (p. ej. traslado de abono al anular)."""
    return max(_to_float(row.get("customer_credit"), 0.0), 0.0)


def _xlsx_border_thin() -> Any:
    from openpyxl.styles import Border, Side

    s = Side(style="thin", color="FF9CA3AF")
    return Border(left=s, right=s, top=s, bottom=s)


def _excel_set_cell(ws: Any, row: int, col: int, value: Any, *, font=None, alignment=None, border=None, fill=None) -> None:
    """Asignación de celda compatible con coordenadas 1-based tipo Excel."""
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill


def _excel_apply_border_block(ws: Any, row_min: int, row_max: int, col_min: int, col_max: int, border: Any) -> None:
    for r in range(row_min, row_max + 1):
        for c in range(col_min, col_max + 1):
            ws.cell(row=r, column=c).border = border


def _citas_filtered_to_excel_bytes(rows: list[dict[str, Any]], *, generated_at: Optional[datetime] = None) -> bytes:
    """Genera .xlsx financiero estilizado: título, encabezados en negrita y tablas demarcadas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    gen_dt = generated_at or datetime.now()
    fecha_etiqueta = gen_dt.strftime("%d/%m/%Y %H:%M")

    datos: list[list[Any]] = []
    for r in rows:
        tot, abo, pend = _financial_row_values(r)
        cred = _customer_credit_value(r)
        nombre = str(r.get("customer_name") or r.get("name") or "").strip()
        artista = _assigned_artist_display_name(r)
        datos.append(
            [
                nombre,
                artista,
                round(tot, 2),
                round(abo, 2),
                round(pend, 2),
                round(cred, 2),
            ]
        )

    headers = [
        "Cliente",
        "Artista / profesional",
        "Valor total (COP)",
        "Abonado (COP)",
        "Pendiente (COP)",
        "Saldo a favor (COP)",
    ]
    ncol = len(headers)

    wb = Workbook()
    bd = _xlsx_border_thin()
    font_title = Font(bold=True, size=14, color="FF111827")
    font_sub = Font(size=10, color="FF4B5563")
    font_header = Font(bold=True, size=11, color="FF111827")
    font_body = Font(size=10, color="FF374151")
    fill_header = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right_num = Alignment(horizontal="right", vertical="center")

    # --- Hoja detalle ---
    ws1 = wb.active
    ws1.title = "Datos financieros"
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws1.cell(1, 1, "Informe financiero — Citas Cherry Ink · Rock City")
    ws1.cell(1, 1).font = font_title
    ws1.cell(1, 1).alignment = align_center

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws1.cell(2, 1, f"Generado: {fecha_etiqueta}")
    ws1.cell(2, 1).font = font_sub
    ws1.cell(2, 1).alignment = align_center

    header_row = 4
    for col in range(1, ncol + 1):
        _excel_set_cell(
            ws1,
            header_row,
            col,
            headers[col - 1],
            font=font_header,
            fill=fill_header,
            alignment=align_left if col <= 2 else align_right_num,
            border=bd,
        )

    row_start_body = header_row + 1
    if datos:
        for i, row_vals in enumerate(datos):
            r = row_start_body + i
            for col in range(1, ncol + 1):
                v = row_vals[col - 1]
                align = align_left if col <= 2 else align_right_num
                _excel_set_cell(ws1, r, col, v, font=font_body, alignment=align, border=bd)
            rmax = r
        _excel_apply_border_block(ws1, header_row, rmax, 1, ncol, bd)
    else:
        ws1.merge_cells(start_row=row_start_body, start_column=1, end_row=row_start_body, end_column=ncol)
        c_msg = ws1.cell(row=row_start_body, column=1, value="Sin filas para los filtros actuales.")
        c_msg.font = font_body
        c_msg.alignment = align_left
        _excel_apply_border_block(ws1, header_row, row_start_body, 1, ncol, bd)

    for col in range(1, ncol + 1):
        letter = get_column_letter(col)
        w = 18
        if col == 1:
            w = 34
        elif col == 2:
            w = 28
        ws1.column_dimensions[letter].width = w

    # --- Resumen ---
    ws2 = wb.create_sheet("Resumen financiero", 1)
    rtot = rabo = rpend = rfav = 0.0
    for rr in rows:
        t, a, p = _financial_row_values(rr)
        rtot += t
        rabo += a
        rpend += p
        rfav += _customer_credit_value(rr)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws2.cell(1, 1, "Resumen financiero — mismos filtros que el panel")
    ws2.cell(1, 1).font = font_title
    ws2.cell(1, 1).alignment = align_center
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws2.cell(2, 1, f"Generado: {fecha_etiqueta}")
    ws2.cell(2, 1).font = font_sub
    ws2.cell(2, 1).alignment = align_center

    resumen_labels = ["Total valor trabajo (COP)", "Total abonado (COP)", "Total pendiente (COP)", "Total saldo a favor (COP)", "Cantidad de citas"]
    resumen_vals: list[Any] = [
        round(rtot, 2),
        round(rabo, 2),
        round(rpend, 2),
        round(rfav, 2),
        len(rows) if rows else 0,
    ]
    hdr_r = 4
    _excel_set_cell(ws2, hdr_r, 1, "Concepto", font=font_header, fill=fill_header, alignment=align_left, border=bd)
    _excel_set_cell(ws2, hdr_r, 2, "Valor", font=font_header, fill=fill_header, alignment=align_right_num, border=bd)
    for i, lab in enumerate(resumen_labels):
        r = hdr_r + 1 + i
        _excel_set_cell(ws2, r, 1, lab, font=font_body, alignment=align_left, border=bd)
        _excel_set_cell(ws2, r, 2, resumen_vals[i], font=font_body, alignment=align_right_num, border=bd)
    _excel_apply_border_block(ws2, hdr_r, hdr_r + len(resumen_labels), 1, 2, bd)
    ws2.column_dimensions["A"].width = 44
    ws2.column_dimensions["B"].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def _parse_date(val: Any) -> date:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str) and val:
        s = val.strip().replace("T", " ")
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    return date(1990, 1, 1)


def _format_appt_when(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip().replace("T", " ")
    if not s:
        return ""
    for c in (s, s[:19], s[:10]):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(c, fmt)
                if fmt == "%Y-%m-%d":
                    return dt.strftime("%d/%m/%Y")
                return dt.strftime("%d/%m/%Y %H:%M")
            except ValueError:
                pass
    return s[:16]


def _time_slot_options() -> List[str]:
    """Franjas cada 30 min entre 08:00 y 20:00 inclusive."""
    slots: List[str] = []
    minutes = 8 * 60
    last = 20 * 60
    while minutes <= last:
        h, mm = divmod(minutes, 60)
        slots.append(f"{h:02d}:{mm:02d}")
        minutes += 30
    return slots


_BOOKING_WORK_KIND_ORDER = ("piercing", "limpieza_piercing", "cambio_piercing", "tatuaje")
_BOOKING_WORK_KIND_META: Dict[str, Dict[str, Any]] = {
    "piercing": {
        "label": "Piercing (colocación)",
        "service_token": "piercing",
        "detail_tag": "[Piercing]",
    },
    "limpieza_piercing": {
        "label": "Limpieza (piercing)",
        "service_token": "piercing",
        "detail_tag": "[Limpieza piercing]",
    },
    "cambio_piercing": {
        "label": "Cambio de piercing",
        "service_token": "piercing",
        "detail_tag": "[Cambio piercing]",
    },
    "tatuaje": {
        "label": "Tatuaje (sesión)",
        "service_token": "tattoo",
        "detail_tag": "[Tatuaje]",
    },
}

_MIN_BOOKING_DURATION_SLOTS = 1
_MAX_BOOKING_DURATION_SLOTS = 16
# Valor mínimo COP para total del trabajo **y** abono inicial al agendar (enteros).
_MIN_APPOINTMENT_TOTAL_COP = 50000.0
# Sufijo en `detail` al crear desde este panel (Streamlit) para saber cuántas franjas de 30 min ocupó la cita.
_AGENDA_SLOTS_DETAIL_PATTERN = re.compile(r"\s*\[agenda_slots:(\d+)\]\s*$", re.IGNORECASE)


def _append_agenda_slots_marker(detail: Optional[str], slots: int) -> str:
    """Añade marcador al final del detalle para calcular ocupación real en agenda."""
    n = max(_MIN_BOOKING_DURATION_SLOTS, min(_MAX_BOOKING_DURATION_SLOTS, int(slots)))
    base = (detail or "").strip()
    return (f"{base} [agenda_slots:{n}]").strip() if base else f"[agenda_slots:{n}]"


def _booking_observations_and_design_for_api() -> str:
    """Combina descripción del diseño y observaciones previas envío a `_service_and_detail_for_work_kind`."""
    dz = str(st.session_state.get("ap_design") or "").strip()
    nt = str(st.session_state.get("ap_det") or "").strip()
    parts: list[str] = []
    if dz:
        parts.append(dz)
    if nt:
        parts.append(nt)
    return "\n".join(parts)


def _booking_duration_slots_from_session() -> int:
    """Franjas de 30 min al agendar (control único; no depende del tipo de trabajo)."""
    raw = st.session_state.get("ap_duration_slots")
    if raw is None:
        return 1
    try:
        n = int(raw)
    except (TypeError, ValueError):
        return 1
    return max(_MIN_BOOKING_DURATION_SLOTS, min(_MAX_BOOKING_DURATION_SLOTS, n))


def _duration_slots_for_existing_appointment(row: dict[str, Any]) -> int:
    """Franjas de 30 min ocupadas por una cita: prioridad al marcador [agenda_slots:N] en detail; si no, heurística."""
    det_full = str(row.get("detail") or "")
    m = _AGENDA_SLOTS_DETAIL_PATTERN.search(det_full)
    if m:
        try:
            return max(
                _MIN_BOOKING_DURATION_SLOTS,
                min(_MAX_BOOKING_DURATION_SLOTS, int(m.group(1))),
            )
        except ValueError:
            pass
    svc = str(row.get("service_type") or row.get("service") or "").strip().lower()
    det = det_full.strip().lower()
    combined = f"{svc} {det}"
    if "limpieza" in det:
        return 1
    if "cambio" in det and "pierc" in combined:
        return 1
    if "tatu" in combined or "tattoo" in svc:
        return 4
    if "pierc" in combined or svc == "piercing":
        return 2
    if "other" in svc or "otro" in svc:
        return 1
    return 2


# Separador explícito entre descripción de diseño y observaciones al editar detalle de cita existente.
_APPT_DESIGN_OBS_DETAIL_SEP = "\n---\n"
_DETAIL_LEADING_BRACKET_TAG = re.compile(r"^\s*\[([^\]\n]{1,64})\]\s*", re.IGNORECASE)


def _appointment_detail_plain_body(detail_full: str) -> str:
    """Quita marcador [agenda_slots:N] y la primera etiqueta tipo [Tatuaje]."""
    d = detail_full or ""
    m = _AGENDA_SLOTS_DETAIL_PATTERN.search(d)
    core = (d[: m.start()] if m else d).strip()
    tm = _DETAIL_LEADING_BRACKET_TAG.match(core)
    if tm:
        core = core[tm.end() :].strip()
    return core


def _split_design_obs_plain(core: str) -> tuple[str, str]:
    if _APPT_DESIGN_OBS_DETAIL_SEP in core:
        a, _, b = core.partition(_APPT_DESIGN_OBS_DETAIL_SEP)
        return a.strip(), b.strip()
    return core.strip(), ""


def _merge_design_obs_plain(design: str, obs: str) -> str:
    dz, nt = design.strip(), obs.strip()
    if dz and nt:
        return f"{dz}{_APPT_DESIGN_OBS_DETAIL_SEP}{nt}"
    return dz or nt


def _work_kind_infer_from_existing_row(row: dict[str, Any]) -> str:
    svc = str(row.get("service_type") or row.get("service") or "").strip().lower()
    det = str(row.get("detail") or "").lower()
    combined = f"{svc} {det}"
    if "limpieza" in det:
        return "limpieza_piercing"
    if "cambio" in det and "pierc" in combined:
        return "cambio_piercing"
    if "tatu" in combined or "tattoo" in svc:
        return "tatuaje"
    if "pierc" in combined or svc == "piercing":
        return "piercing"
    return "piercing"


def _rebuild_detail_for_patch(
    row: dict[str, Any],
    design: str,
    obs: str,
    *,
    agenda_slots_override: Optional[int] = None,
) -> str:
    slots = (
        int(agenda_slots_override)
        if agenda_slots_override is not None
        else _duration_slots_for_existing_appointment(row)
    )
    slots = max(_MIN_BOOKING_DURATION_SLOTS, min(_MAX_BOOKING_DURATION_SLOTS, slots))
    wk = _work_kind_infer_from_existing_row(row)
    merged = _merge_design_obs_plain(design, obs)
    _, detail_for_api = _service_and_detail_for_work_kind(wk, merged)
    return _append_agenda_slots_marker(detail_for_api, slots)


def _appointment_last_start_slot(start_hm: str, dur_slots: int, slot_opts: list[str]) -> str:
    try:
        si = slot_opts.index(start_hm)
    except ValueError:
        return start_hm
    last_i = min(si + max(int(dur_slots), 1) - 1, len(slot_opts) - 1)
    return slot_opts[last_i]


def _duration_slots_from_start_last(start_hm: str, last_slot_hm: str, slot_opts: list[str]) -> int:
    try:
        si = slot_opts.index(start_hm)
        ei = slot_opts.index(last_slot_hm)
    except ValueError:
        return _MIN_BOOKING_DURATION_SLOTS
    if ei < si:
        return _MIN_BOOKING_DURATION_SLOTS
    return max(_MIN_BOOKING_DURATION_SLOTS, min(_MAX_BOOKING_DURATION_SLOTS, ei - si + 1))


def _format_appt_created_at(val: Any) -> str:
    if val is None:
        return "—"
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    s = str(val).strip()
    if len(s) >= 16:
        return s[:16].replace("T", " ")
    return s or "—"


def _appointments_same_day_raw(items: list[dict[str, Any]], day: date) -> list[dict[str, Any]]:
    """Citas de ese día usando la lista completa de API (sin filtrar por nombre), para no solapar huecos."""
    out: list[dict[str, Any]] = []
    for row in items:
        try:
            d = _parse_date(row.get("appointment_date", row.get("date")))
        except (TypeError, ValueError):
            continue
        if d != day:
            continue
        out.append(row)
    return out


def _consume_cal_appt_query_param() -> None:
    """Abre el diálogo de una cita desde enlaces dentro de la rejilla (`?cal_appt_id=`)."""
    raw = st.query_params.get("cal_appt_id")
    if raw is None:
        return
    try:
        aid = int(str(raw).strip())
        if aid > 0:
            st.session_state["_cal_focus_appt_id"] = aid
            st.session_state.pop("_cal_overflow_day", None)
    except ValueError:
        pass
    st.query_params.pop("cal_appt_id", None)


def _twg_sess_appt_open_button_key(monday: date, aid: int) -> str:
    """Clave estable para `st.button` oculto que abre una cita desde la rejilla semanal."""
    return f"twg_sess_appt_open_{monday.isoformat()}_{aid}"


def _week_grid_visible_appt_ids(
    monday: date,
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
) -> list[int]:
    """IDs de citas visibles en la semana (orden por día y aparición)."""
    out: list[int] = []
    seen: set[int] = set()
    for i in range(7):
        d = monday + timedelta(days=i)
        for row in buckets.get((d.year, d.month, d.day), []):
            aid = int(row.get("id", 0) or 0)
            if aid <= 0 or aid in seen:
                continue
            seen.add(aid)
            out.append(aid)
    return out


def _render_week_grid_hidden_appt_open_buttons(
    monday: date,
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
) -> None:
    """Botones Streamlit invisibles; el clic en la pastilla HTML delega aquí vía JS (sin navegar por URL)."""
    ids = _week_grid_visible_appt_ids(monday, buckets)
    if not ids:
        return
    keys = [_twg_sess_appt_open_button_key(monday, aid) for aid in ids]
    css_chunks = [
        "section.main button."
        + html_mod.escape(f"st-key-{k}")
        + "{position:absolute!important;left:-9999px!important;width:1px!important;"
        "height:1px!important;opacity:0!important;margin:0!important;padding:0!important;}"
        for k in keys
    ]
    st.markdown("<style>" + "\n".join(css_chunks) + "</style>", unsafe_allow_html=True)
    for aid in ids:
        bk = _twg_sess_appt_open_button_key(monday, aid)
        if st.button(
            "\u200b",
            key=bk,
            help=f"Abrir cita #{aid} (rejilla semanal)",
            type="tertiary",
            width=1,
        ):
            st.session_state["_cal_focus_appt_id"] = aid
            st.session_state.pop("_cal_overflow_day", None)
            st.rerun()


def _inject_week_grid_appt_click_bridge(monday_iso: str) -> None:
    """Engancha clics en `.twg-appt-link` y dispara el `st.button` oculto correspondiente."""
    mon_lit = json.dumps(monday_iso)
    inner_js = f"""
(function () {{
  var NS = "__twgApptOpenBridge_v2";
  var mondayIso = {mon_lit};

  function appDocument() {{
    try {{
      var t = window.top;
      if (t && t.document && t.document.querySelector('[data-testid="stApp"]')) return t.document;
    }} catch (e0) {{}}
    var x = window;
    for (var i = 0; i < 12; i++) {{
      try {{
        if (!x || !x.document) break;
        var d = x.document;
        if (d.querySelector('[data-testid="stApp"]')) return d;
        if (x.parent === x) break;
        x = x.parent;
      }} catch (e) {{
        break;
      }}
    }}
    try {{
      return window.top.document;
    }} catch (e2) {{
      return document;
    }}
  }}

  function prevTuple(tupleVal) {{
    if (
      !tupleVal ||
      !tupleVal.shell ||
      !tupleVal.onClick ||
      typeof tupleVal.shell.removeEventListener !== "function"
    ) {{
      return;
    }}
    try {{
      tupleVal.shell.removeEventListener("click", tupleVal.onClick, true);
    }} catch (e0) {{}}
  }}

  function mount() {{
    var prev = window[NS];
    if (prev && typeof prev === "object") {{
      prevTuple(prev);
    }}

    function findWidgetButton(doc, aid) {{
      var cls = "st-key-twg_sess_appt_open_" + mondayIso + "_" + aid;
      var esc =
        typeof CSS !== "undefined" && typeof CSS.escape === "function"
          ? CSS.escape(cls)
          : cls;
      try {{
        var hosts = doc.querySelectorAll("." + esc);
        for (var i = 0; i < hosts.length; i++) {{
          var h = hosts[i];
          var btn = h.tagName === "BUTTON" ? h : h.querySelector("button");
          if (btn) return btn;
        }}
      }} catch (e1) {{}}
      return null;
    }}

    function onClick(ev) {{
      var t = ev.target;
      if (!t || typeof t.closest !== "function") return;
      var pill = t.closest("button.twg-appt-link[data-cal-appt-id]");
      if (!pill) return;
      var raw = pill.getAttribute("data-cal-appt-id");
      if (!raw) return;
      ev.preventDefault();
      ev.stopPropagation();
      var aid = parseInt(raw, 10);
      if (!aid || aid <= 0) return;
      var doc = appDocument();
      var wb = findWidgetButton(doc, aid);
      if (wb) {{
        try {{
          wb.click();
        }} catch (e2) {{}}
      }}
    }}

    function bind() {{
      var doc = appDocument();
      var shell = doc.querySelector(".twg-shell");
      if (!shell) return;
      shell.removeEventListener("click", onClick, true);
      shell.addEventListener("click", onClick, true);
      window[NS] = {{ mondayIso: mondayIso, shell: shell, onClick: onClick }};
    }}

    bind();
    setTimeout(bind, 0);
    setTimeout(bind, 120);
    setTimeout(bind, 400);
  }}

  mount();
}})();
"""
    outer = (
        "<script>\n"
        "(function () {\n"
        '  function appDocument() {\n'
        "    try {\n"
        '      var t = window.top;\n'
        '      if (t && t.document && t.document.querySelector(\'[data-testid="stApp"]\')) return t.document;\n'
        "    } catch (e0) {}\n"
        "    var x = window;\n"
        "    for (var i = 0; i < 12; i++) {\n"
        "      try {\n"
        "        if (!x || !x.document) break;\n"
        "        var d = x.document;\n"
        '        if (d.querySelector(\'[data-testid="stApp"]\')) return d;\n'
        "        if (x.parent === x) break;\n"
        "        x = x.parent;\n"
        "      } catch (e) {\n"
        "        break;\n"
        "      }\n"
        "    }\n"
        "    try {\n"
        "      return window.top.document;\n"
        "    } catch (e2) {\n"
        "      return document;\n"
        "    }\n"
        "  }\n"
        "  function injectScript(doc, text) {\n"
        "    try {\n"
        '      var s = doc.createElement("script");\n'
        "      s.textContent = text;\n"
        "      (doc.head || doc.documentElement).appendChild(s);\n"
        "      s.remove();\n"
        "    } catch (e) {}\n"
        "  }\n"
        f"  injectScript(appDocument(), {json.dumps(inner_js)});\n"
        "})();\n"
        "</script>"
    )
    components.html(outer, height=0, width=0)


def _calendar_month_footer_strip_html(
    y: int,
    m: int,
    d: int,
    *,
    disabled_footer: bool,
    title_esc: str,
) -> str:
    """Pie de celda mensual tipo cabecera: día visible + botón HTML que delega en `st.button` oculto."""
    strip_cls = "cal-cell-footer-strip"
    if disabled_footer:
        strip_cls += " cal-footer-strip-disabled"
    bd = " disabled" if disabled_footer else ""
    return (
        f'<div class="{strip_cls}" role="presentation">'
        f'<span class="cal-footer-daynum">{html_mod.escape(str(d))}</span>'
        f"<button type=\"button\" class=\"cal-footer-book\" "
        f'data-cal-y="{int(y)}" data-cal-m="{int(m)}" data-cal-d="{int(d)}" '
        f'title="{title_esc}" aria-label="{title_esc}"{bd}>'
        "+ Agregar cita</button>"
        "</div>"
    )


def _render_calendar_month_hidden_book_widgets(y: int, m: int, weeks_grid: list[list[int]]) -> None:
    """Botones ocultos (uno por día) para poder agendar desde el pie HTML mediante JS."""
    allow_book_role = not _panel_is_technician_role()
    booked: list[int] = []
    for row in weeks_grid:
        for d in row:
            if d > 0:
                booked.append(d)
    if not booked:
        return
    css_chunks = [
        "section.main button[data-testid^=\"baseButton\"][class*=\"cal_main_day_\"] "
        "{position:absolute!important;left:-9999px!important;width:1px!important;"
        "height:1px!important;opacity:0!important;margin:0!important;padding:0!important;}",
        'section.main div[data-testid="element-container"]:has(button[class*="cal_main_day_"]) '
        "{min-height:0!important;max-height:0!important;margin:0!important;padding:0!important;"
        "border:none!important;overflow:hidden!important;}",
        "section.main div[data-testid=\"stHorizontalBlock\"]:has(button[class*=\"cal_main_day_\"]) "
        "{min-height:0!important;margin:0!important;padding:0!important;gap:0!important;}",
        "section.main div[data-testid=\"stHorizontalBlock\"]:has(button[class*=\"cal_main_day_\"]) "
        "> div[data-testid=\"column\"] {min-height:0!important;padding:0!important;flex:0 0 auto!important;}",
        "section.main div[data-testid=\"stVerticalBlock\"]:has(div[data-testid=\"stHorizontalBlock\"] button[class*=\"cal_main_day_\"]) "
        "{gap:0!important;}",
    ]
    st.markdown("<style>" + "\n".join(css_chunks) + "</style>", unsafe_allow_html=True)
    today_d = date.today()
    # En filas de 7 como la rejilla mensual para no encadenar docenas de filas Streamlit ocultando el scroll útil del calendario.
    chunk = 7
    for i in range(0, len(booked), chunk):
        row_days = booked[i : i + chunk]
        row_cols = st.columns(len(row_days), gap=None)
        for ci, d in enumerate(row_days):
            with row_cols[ci]:
                picked_cell = date(y, m, d)
                bk = f"cal_main_day_{y}_{m}_{d}"
                is_past = picked_cell < today_d
                disabled = is_past or not allow_book_role
                book_help = (
                    "Los tatuadores y perforadores no pueden agendar desde el calendario"
                    if not allow_book_role
                    else (
                        "No se pueden agendar citas en fechas pasadas"
                        if is_past
                        else f"Agendar cita · {picked_cell.strftime('%d/%m/%Y')}"
                    )
                )
                if st.button(
                    "\u200b",
                    key=bk,
                    type="tertiary",
                    width=1,
                    disabled=disabled,
                    help=book_help,
                ):
                    _clear_calendar_dialog_focus()
                    _pop_booking_document_session()
                    st.session_state["ap_ad"] = date(y, m, d)
                    st.session_state["_ap_dlg"] = "create"
                    st.rerun()


def _inject_calendar_month_footer_bridge() -> None:
    """Clicks en `.cal-footer-book` disparan el `st.button` oculto `cal_main_day_…`."""
    inner_js = f"""
(function () {{
  var NS = "__calFooterBook_v1";

  function appDocument() {{
    try {{
      var t = window.top;
      if (t && t.document && t.document.querySelector('[data-testid="stApp"]')) return t.document;
    }} catch (e0) {{}}
    var x = window;
    for (var i = 0; i < 12; i++) {{
      try {{
        if (!x || !x.document) break;
        var d = x.document;
        if (d.querySelector('[data-testid="stApp"]')) return d;
        if (x.parent === x) break;
        x = x.parent;
      }} catch (e) {{
        break;
      }}
    }}
    try {{
      return window.top.document;
    }} catch (e2) {{
      return document;
    }}
  }}

  function prevTuple(tupleVal) {{
    if (!tupleVal || !tupleVal.shell || !tupleVal.onClick) return;
    try {{
      tupleVal.shell.removeEventListener("click", tupleVal.onClick, true);
    }} catch (e0) {{}}
  }}

  function mount() {{
    var prev = window[NS];
    if (prev && typeof prev === "object") prevTuple(prev);

    function findBookButton(doc, yy, mm, dd) {{
      /* Misma convención que la rejilla semanal (clase st-key- + subKey) */
      var subKey = "cal_main_day_" + yy + "_" + mm + "_" + dd;
      var candidates = ["st-key-" + subKey, subKey];
      for (var c = 0; c < candidates.length; c++) {{
        var cls = candidates[c];
        var esc =
          typeof CSS !== "undefined" && typeof CSS.escape === "function"
            ? CSS.escape(cls)
            : cls;
        try {{
          var hosts = doc.querySelectorAll("." + esc);
          for (var i = 0; i < hosts.length; i++) {{
            var h = hosts[i];
            var btn =
              h.tagName === "BUTTON"
                ? h
                : h.querySelector('button[data-testid^="baseButton"]') || h.querySelector("button");
            if (btn && !btn.disabled) return btn;
          }}
        }} catch (eCls) {{}}
      }}
      try {{
        var hits = doc.querySelectorAll('[class*="' + subKey + '"]');
        for (var j = 0; j < hits.length; j++) {{
          var node = hits[j];
          var btn =
            node.tagName === "BUTTON"
              ? node
              : node.querySelector('button[data-testid^="baseButton"]');
          if (btn && !btn.disabled) return btn;
        }}
      }} catch (eSub) {{}}
      return null;
    }}

    function onClick(ev) {{
      var t = ev.target;
      if (!t || typeof t.closest !== "function") return;
      var el = t.closest("button.cal-footer-book");
      if (!el || el.closest(".cal-footer-strip-disabled")) return;
      if (el.disabled) return;
      ev.preventDefault();
      ev.stopPropagation();
      var cy = parseInt(el.getAttribute("data-cal-y") || "", 10);
      var cm = parseInt(el.getAttribute("data-cal-m") || "", 10);
      var cd = parseInt(el.getAttribute("data-cal-d") || "", 10);
      if (!cd || cd <= 0) return;
      var doc = appDocument();
      var wb = findBookButton(doc, cy, cm, cd);
      if (wb) {{
        try {{
          wb.click();
        }} catch (e2) {{
          try {{
            wb.dispatchEvent(new MouseEvent("click", {{ bubbles: true, cancelable: true, view: window }}));
          }} catch (e3) {{}}
        }}
      }}
    }}

    function bind() {{
      var doc = appDocument();
      var shell =
        doc.querySelector('[data-testid="stMain"]') ||
        doc.querySelector("section.main") ||
        doc.body;
      if (!shell) return;
      shell.removeEventListener("click", onClick, true);
      shell.addEventListener("click", onClick, true);
      window[NS] = {{ shell: shell, onClick: onClick }};
    }}

    bind();
    setTimeout(bind, 0);
    setTimeout(bind, 120);
    setTimeout(bind, 400);
  }}

  mount();
}})();
"""

    outer = (
        "<script>\n"
        "(function () {\n"
        '  function appDocument() {\n'
        "    try {\n"
        '      var t = window.top;\n'
        '      if (t && t.document && t.document.querySelector(\'[data-testid="stApp"]\')) return t.document;\n'
        "    } catch (e0) {}\n"
        "    var x = window;\n"
        "    for (var i = 0; i < 12; i++) {\n"
        "      try {\n"
        "        if (!x || !x.document) break;\n"
        "        var d = x.document;\n"
        '        if (d.querySelector(\'[data-testid="stApp"]\')) return d;\n'
        "        if (x.parent === x) break;\n"
        "        x = x.parent;\n"
        "      } catch (e) {\n"
        "        break;\n"
        "      }\n"
        "    }\n"
        "    try {\n"
        "      return window.top.document;\n"
        "    } catch (e2) {\n"
        "      return document;\n"
        "    }\n"
        "  }\n"
        "  function injectScript(doc, text) {\n"
        "    try {\n"
        '      var s = doc.createElement("script");\n'
        "      s.textContent = text;\n"
        "      (doc.head || doc.documentElement).appendChild(s);\n"
        "      s.remove();\n"
        "    } catch (e) {}\n"
        "  }\n"
        f"  injectScript(appDocument(), {json.dumps(inner_js)});\n"
        "})();\n"
        "</script>"
    )
    components.html(outer, height=0, width=0)


def _busy_slot_indices_for_day(day_rows: list[dict[str, Any]], slot_list: list[str]) -> set[int]:
    busy: set[int] = set()
    n = len(slot_list)
    for row in day_rows:
        if str(row.get("status") or "").strip().lower() == "cancelada":
            continue
        hm = _appt_time_hm(row.get("appointment_date", row.get("date")))
        if hm == "—":
            continue
        try:
            start_idx = slot_list.index(hm)
        except ValueError:
            continue
        dur = _duration_slots_for_existing_appointment(row)
        for j in range(start_idx, min(start_idx + dur, n)):
            busy.add(j)
    return busy


def _available_start_slots(slot_list: list[str], need_slots: int, busy: set[int]) -> list[str]:
    n = len(slot_list)
    out: list[str] = []
    for i in range(n):
        if i + need_slots > n:
            break
        if any(j in busy for j in range(i, i + need_slots)):
            continue
        out.append(slot_list[i])
    return out


def _service_and_detail_for_work_kind(kind: str, user_detail: str) -> tuple[str, Optional[str]]:
    meta = _BOOKING_WORK_KIND_META.get(kind) or _BOOKING_WORK_KIND_META["piercing"]
    svc = resolve_service_type(meta["service_token"])
    tag = meta["detail_tag"]
    extra = (user_detail or "").strip()
    if extra:
        return svc, f"{tag} {extra}".strip()
    return svc, tag


def _combine_appointment_datetime(d: date, slot_hm: str) -> str:
    slot_hm = (slot_hm or "09:00").strip()
    parts = slot_hm.split(":")
    h = int(parts[0])
    m = int(parts[1]) if len(parts) > 1 else 0
    return f"{d.strftime('%Y-%m-%d')} {h:02d}:{m:02d}:00"


def _parse_existing_slot(val: Any) -> tuple[date, str]:
    d = _parse_date(val)
    s = str(val or "").strip().replace("T", " ")
    opts = _time_slot_options()
    if len(s) >= 16:
        chunk = s[11:16]
        if chunk in opts:
            return d, chunk
    return d, "09:00" if "09:00" in opts else opts[0]


def _appt_time_hm(val: Any) -> str:
    """Hora HH:MM para listados compactos; '—' si solo hay fecha."""
    if val is None:
        return "—"
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, date) and not isinstance(val, datetime):
        return "—"
    s = str(val).strip().replace("T", " ")
    if not s:
        return "—"
    for chunk, fmt in ((s[:19], "%Y-%m-%d %H:%M:%S"), (s[:16], "%Y-%m-%d %H:%M")):
        try:
            return datetime.strptime(chunk, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return "—"


def _monday_of_week(d: date) -> date:
    """Lunes como primer día de la semana (ISO, weekday() 0=lun)."""
    return d - timedelta(days=d.weekday())


def _sync_week_monday_for_agenda_context() -> None:
    """Deja `_ap_week_monday` coherente con el mes abierto en el calendario (o con hoy si es el mismo mes)."""
    ym_raw = st.session_state.get("_ap_cal_ym")
    if isinstance(ym_raw, (list, tuple)) and len(ym_raw) >= 2:
        try:
            y, m = int(ym_raw[0]), int(ym_raw[1])
        except (TypeError, ValueError):
            td = date.today()
            y, m = td.year, td.month
    else:
        td = date.today()
        y, m = td.year, td.month
    today_d = date.today()
    anchor = today_d if (today_d.year == y and today_d.month == m) else date(y, m, 1)
    st.session_state["_ap_week_monday"] = _monday_of_week(anchor).isoformat()


# Altura en px de cada franja de 30 min en la vista semanal tipo Teams/Outlook.
_TWGS_SLOT_PX = 22


def _week_grid_parse_segments(
    day_rows: list[dict[str, Any]],
    slot_list: list[str],
) -> list[tuple[dict[str, Any], int, int]]:
    """Citas posicionables en la rejilla: (row, start_idx, dur_slots_vis)."""
    n = len(slot_list)
    raw: list[tuple[dict[str, Any], int, int]] = []
    for row in day_rows:
        hm = _appt_time_hm(row.get("appointment_date", row.get("date")))
        if hm == "—":
            continue
        try:
            start_idx = slot_list.index(hm)
        except ValueError:
            continue
        dur = _duration_slots_for_existing_appointment(row)
        end_idx = min(start_idx + dur, n)
        vis = max(0, end_idx - start_idx)
        if vis <= 0:
            continue
        raw.append((row, start_idx, vis))
    raw.sort(key=lambda x: (x[1], -x[2]))
    return raw


def _week_grid_assign_lanes(
    raw: list[tuple[dict[str, Any], int, int]],
) -> list[tuple[dict[str, Any], int, int, int, int]]:
    """Asigna carril para solapes (greedy). Devuelve (row, start_idx, vis, lane_i, n_lanes)."""
    lanes_end: list[int] = []
    assignments: list[tuple[dict[str, Any], int, int, int]] = []
    for row, start_idx, vis in raw:
        end_idx = start_idx + vis
        lane_i = -1
        for li, le in enumerate(lanes_end):
            if start_idx >= le:
                lane_i = li
                lanes_end[li] = end_idx
                break
        if lane_i < 0:
            lane_i = len(lanes_end)
            lanes_end.append(end_idx)
        assignments.append((row, start_idx, vis, lane_i))
    n_lanes = max(1, len(lanes_end))
    return [(r, s, v, l, n_lanes) for r, s, v, l in assignments]


def _week_grid_day_column_html(
    d: date,
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
    counts_by_client: dict[str, int],
    *,
    slot_list: list[str],
    slot_px: int,
    today: date,
) -> str:
    """Columna de un día con líneas de tiempo y bloques absolutos."""
    n_slots = len(slot_list)
    total_h = n_slots * slot_px
    day_rows = list(buckets.get((d.year, d.month, d.day), []))
    raw = _week_grid_parse_segments(day_rows, slot_list)
    placed = _week_grid_assign_lanes(raw)
    today_cls = " twg-col-today" if d == today else ""
    parts: list[str] = [f'<div class="twg-col{today_cls}" style="height:{total_h}px">']
    for si in range(n_slots):
        parts.append(
            f'<div class="twg-slot-line" style="top:{si * slot_px}px;height:{slot_px}px"></div>'
        )
    for row, start_idx, vis, lane_i, n_lanes in placed:
        cls_pill = _client_pill_class(row, counts_by_client)
        top = start_idx * slot_px
        height = max(vis * slot_px - 2, 14)
        pct_w = 100.0 / n_lanes
        left = lane_i * pct_w
        nm = str(row.get("customer_name") or row.get("name") or "").strip() or "—"
        nm_esc = html_mod.escape(nm)
        hm = _appt_time_hm(row.get("appointment_date", row.get("date")))
        staff = _assigned_staff_label(row)
        tit = html_mod.escape(f"{hm} · {nm}" + (f" · {staff}" if staff != "—" else ""))
        st_cl = str(row.get("status") or "").strip().lower()
        soft_cls = " twg-cancelada-soft" if st_cl == "cancelada" else ""
        ap_id = int(row.get("id", 0) or 0)
        # Una sola franja (= 30 min): el alto es ~20px; sin layout horizontal no cabe nombre + botón.
        compact = vis <= 1
        link_html = ""
        if ap_id > 0:
            if compact:
                link_html = (
                    f'<button type="button" class="twg-appt-link twg-appt-link--compact" '
                    f'data-cal-appt-id="{ap_id}" aria-label="Ver cita">'
                    '<span class="twg-appt-link-play" aria-hidden="true">▶</span>'
                    "</button>"
                )
            else:
                link_html = (
                    f'<button type="button" class="twg-appt-link" data-cal-appt-id="{ap_id}" '
                    'aria-label="Ver cita">'
                    '<span class="twg-appt-link-play" aria-hidden="true">▶</span>'
                    "<span>Ver cita</span>"
                    "</button>"
                )
        cls_appt_extra = " twg-appt--compact" if compact else ""
        if compact:
            body = (
                f'<span class="twg-appt-time-compact">{html_mod.escape(hm)}</span>'
                f'<span class="twg-appt-client-compact">{nm_esc}</span>'
            )
        else:
            body = (
                '<div class="twg-appt-body-stack">'
                f'<span class="twg-appt-head-time">{html_mod.escape(hm)}</span>'
                f'<span class="twg-appt-client">{nm_esc}</span>'
                "</div>"
            )
        geo = (
            f"top:{top}px;height:{height}px;left:{left}%;width:calc({pct_w}% - 3px);margin-left:1px"
        )
        parts.append(
            f'<div class="twg-appt twg-{cls_pill}{cls_appt_extra}{soft_cls}" '
            f'title="{tit}" style="{geo}">{body}{link_html}</div>'
        )
    parts.append("</div>")
    return "".join(parts)


def _week_grid_html_for_week(
    monday: date,
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
    counts_by_client: dict[str, int],
    *,
    slot_list: list[str],
    slot_px: int,
) -> str:
    """HTML de la rejilla semanal completa (cabecera + cuerpo)."""
    today = date.today()
    weekdays_short = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    parts: list[str] = ['<div class="twg-shell"><div class="twg-scroll-x"><div class="twg-grid-wrap">']

    parts.append('<div class="twg-head">')
    parts.append('<div class="twg-h-spacer"></div>')
    parts.append('<div class="twg-h-days">')
    for i in range(7):
        d = monday + timedelta(days=i)
        today_cls = " twg-h-today" if d == today else ""
        parts.append(
            f'<div class="twg-h-day{today_cls}"><div class="twg-h-wd">{weekdays_short[i]}</div>'
            f'<div class="twg-h-num">{d.day}</div></div>'
        )
    parts.append("</div></div>")

    parts.append('<div class="twg-body">')
    parts.append('<div class="twg-times">')
    for si, hm in enumerate(slot_list):
        is_hour = hm.endswith(":00")
        cls = "twg-tick-major" if is_hour else "twg-tick-minor"
        label = hm if is_hour else ""
        parts.append(
            f'<div class="twg-tick {cls}" style="height:{slot_px}px"><span>{label}</span></div>'
        )
    parts.append("</div>")
    parts.append('<div class="twg-day-columns">')
    for i in range(7):
        d = monday + timedelta(days=i)
        parts.append(
            _week_grid_day_column_html(
                d,
                buckets,
                counts_by_client,
                slot_list=slot_list,
                slot_px=slot_px,
                today=today,
            )
        )
    parts.append("</div></div>")
    parts.append("</div></div></div>")
    return "".join(parts)


def _render_week_schedule_grid(
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
    counts_by_client: dict[str, int],
) -> None:
    """Vista semanal con columnas por día y bloques proporcionales a la duración (30 min/slot)."""
    anchor_key = "_ap_week_monday"
    if anchor_key not in st.session_state:
        st.session_state[anchor_key] = _monday_of_week(date.today()).isoformat()
    try:
        monday = date.fromisoformat(str(st.session_state[anchor_key]))
    except ValueError:
        monday = _monday_of_week(date.today())
        st.session_state[anchor_key] = monday.isoformat()

    st.markdown("##### Agenda semanal (rejilla horaria)")
    st.caption(
        "Estilo **Outlook / Teams**: columnas por día, franjas de **30 min** (**08:00–20:00**). "
        "En cada bloque, **Ver cita** (▶) abre el detalle en un **panel emergente** sin cambiar de página; "
        "también puedes usar **Lista** para ver todas las citas del día."
    )

    b1, b2, b3, b4 = st.columns([1.1, 1.1, 3.2, 1.1])
    with b1:
        if st.button("◀ Semana", key="twg_prev_week"):
            _clear_calendar_dialog_focus()
            st.session_state[anchor_key] = (monday - timedelta(days=7)).isoformat()
            st.rerun()
    with b2:
        if st.button("Hoy", key="twg_today_week"):
            _clear_calendar_dialog_focus()
            st.session_state[anchor_key] = _monday_of_week(date.today()).isoformat()
            st.rerun()
    with b3:
        sunday = monday + timedelta(days=6)
        span_lbl = f"{monday.strftime('%d/%m')} – {sunday.strftime('%d/%m/%Y')}"
        st.markdown(
            f"<div style='text-align:center;font-weight:600;padding-top:0.35rem'>{html_mod.escape(span_lbl)}</div>",
            unsafe_allow_html=True,
        )
    with b4:
        if st.button("Semana ▶", key="twg_next_week"):
            _clear_calendar_dialog_focus()
            st.session_state[anchor_key] = (monday + timedelta(days=7)).isoformat()
            st.rerun()

    slot_list = _time_slot_options()
    st.markdown(
        _week_grid_html_for_week(
            monday,
            buckets,
            counts_by_client,
            slot_list=slot_list,
            slot_px=_TWGS_SLOT_PX,
        ),
        unsafe_allow_html=True,
    )
    _render_week_grid_hidden_appt_open_buttons(monday, buckets)
    _inject_week_grid_appt_click_bridge(monday.isoformat())

    st.markdown("**Acciones por día**")
    wd_short = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    cols = st.columns(7)
    allow_book = not _panel_is_technician_role()
    for i in range(7):
        d = monday + timedelta(days=i)
        day_rows = list(buckets.get((d.year, d.month, d.day), []))
        with cols[i]:
            st.caption(f"{wd_short[i]} **{d.day}** · {len(day_rows)} cita(s)")
            if st.button(
                "Lista",
                key=f"twg_day_list_{d.isoformat()}",
                use_container_width=True,
                disabled=len(day_rows) == 0,
                help="Ver todas las citas del día en un solo panel",
            ):
                st.session_state.pop("_cal_focus_appt_id", None)
                st.session_state["_cal_overflow_day"] = (d.year, d.month, d.day)
                st.rerun()
            past = d < date.today()
            if st.button(
                "Agendar",
                key=f"twg_day_book_{d.isoformat()}",
                use_container_width=True,
                disabled=past or not allow_book,
                help=(
                    "Los tatuadores y perforadores no pueden agendar desde aquí"
                    if not allow_book
                    else ("No se pueden agendar fechas pasadas" if past else "Nueva cita en este día")
                ),
            ):
                _clear_calendar_dialog_focus()
                _pop_booking_document_session()
                st.session_state["ap_ad"] = d
                st.session_state["_ap_dlg"] = "create"
                st.rerun()


def _appointments_by_day_sorted(items: list[dict[str, Any]]) -> dict[tuple[int, int, int], list[dict[str, Any]]]:
    """Citas agrupadas por día local, ordenadas por hora de inicio."""
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]] = {}

    def sort_key(r: dict[str, Any]) -> tuple[int, int]:
        raw = r.get("appointment_date", r.get("date"))
        if isinstance(raw, datetime):
            return (raw.hour, raw.minute)
        s = str(raw or "").strip().replace("T", " ")
        for chunk, fmt in ((s[:19], "%Y-%m-%d %H:%M:%S"), (s[:16], "%Y-%m-%d %H:%M")):
            try:
                dt = datetime.strptime(chunk, fmt)
                return (dt.hour, dt.minute)
            except ValueError:
                pass
        return (99, 99)

    for row in items:
        try:
            d = _parse_date(row.get("appointment_date", row.get("date")))
        except (TypeError, ValueError):
            continue
        key = (d.year, d.month, d.day)
        buckets.setdefault(key, []).append(row)
    for appts in buckets.values():
        appts.sort(key=sort_key)
    return buckets


def _appointment_counts_by_client(items: list[dict[str, Any]]) -> dict[str, int]:
    """Total de citas por cliente en todo el histórico cargado (lista API)."""
    counts: dict[str, int] = {}
    for row in items:
        k = _client_history_key(row)
        counts[k] = counts.get(k, 0) + 1
    return counts


def _calendar_overflow_row_html(row: dict[str, Any], counts_by_client: dict[str, int]) -> str:
    """Línea para el diálogo de citas extra: hora + tipo de servicio + nombre con pill de cliente + total."""
    hm = _appt_time_hm(row.get("appointment_date", row.get("date")))
    st_cl = str(row.get("status") or "").strip().lower()
    muted = " cal-overflow-row--muted" if st_cl == "cancelada" else ""
    svc_flag = _service_type_flag_html(row)
    pill = _customer_name_pill_html(row, counts_by_client)
    staff_s = _assigned_artist_display_name(row)
    staff_el = ""
    if staff_s != "Sin asignar":
        staff_el = (
            '<span class="cal-overflow-artist-dash"> · Artista: '
            f"{html_mod.escape(staff_s)}</span>"
        )
    total_amt, _, _ = _financial_row_values(row)
    total_box = (
        '<span class="cal-overflow-total-wrap"><span class="cal-overflow-total-chip">'
        '<span class="cal-overflow-total-label">Total servicio</span> · '
        f'{html_mod.escape(_format_cop(total_amt))}</span></span>'
    )
    fire = ""
    if bool(row.get("contract_pending_artist_signature")):
        fire = '<span class="cal-overflow-fire-pending">Firma profesional pendiente</span>'
    left_cluster = (
        '<span class="cal-overflow-left">'
        f'<span class="cal-overflow-hm">{html_mod.escape(hm)}</span><span>·</span>'
        f"{svc_flag}<span>·</span>{pill}{staff_el}"
        f"{fire}</span>"
    )
    return f'<div class="cal-overflow-row{muted}">{left_cluster}{total_box}</div>'


def _calendar_cell_customer_label(full_name: str, *, long_from_len: int = 18) -> str:
    """Texto compacto en celda del calendario: nombre completo si es corto; si no, solo el primer nombre."""
    nm = (full_name or "").strip() or "—"
    if nm == "—" or len(nm) <= long_from_len:
        return nm
    parts = nm.split()
    if not parts:
        return nm[: long_from_len - 1] + "…"
    first = parts[0]
    if len(first) > long_from_len:
        return first[: long_from_len - 1] + "…"
    return first


def _group_calendar_day_rows_by_assigned_staff(
    day_rows: list[dict[str, Any]],
) -> list[tuple[str, list[dict[str, Any]]]]:
    """Agrupa las citas del día por profesional asignado (orden de grupos = primera aparición en la línea de tiempo del día)."""
    order_keys: list[Any] = []
    buckets: dict[Any, list[dict[str, Any]]] = {}
    labels: dict[Any, str] = {}

    for r in day_rows:
        raw = r.get("assigned_panel_user_id")
        key: Any
        try:
            key = int(raw) if raw is not None and str(raw).strip() != "" else None
        except (TypeError, ValueError):
            key = None
        if key is None:
            key = "__unassigned__"
            lab = "Sin asignar"
        else:
            lab = _assigned_artist_display_name(r)
        if key not in buckets:
            buckets[key] = []
            order_keys.append(key)
            labels[key] = lab
        buckets[key].append(r)

    return [(labels[k], buckets[k]) for k in order_keys]


def _calendar_day_inner_body_html(
    day_rows: list[dict[str, Any]],
    counts_by_client: dict[str, int],
    *,
    team_layout: bool,
) -> tuple[str, int]:
    """HTML dentro de cal-day-inner: todas las citas del día; el alto fijo de la celda usa scroll."""
    if not day_rows:
        return "<div class='cal-day-empty'>—</div>", 0

    chunks: list[str] = []
    shown = 0

    if team_layout:
        for label, rows_g in _group_calendar_day_rows_by_assigned_staff(day_rows):
            line_htmls: list[str] = []
            for r in rows_g:
                line_htmls.append(_calendar_appt_line_html(r, counts_by_client))
                shown += 1
            if line_htmls:
                chunks.append(
                    "<div class='cal-team-block'>"
                    f"<div class='cal-team-label'>{html_mod.escape(label)}</div>"
                    "<div class='cal-team-lines'>"
                    f"{''.join(line_htmls)}"
                    "</div></div>"
                )
    else:
        for r in day_rows:
            chunks.append(_calendar_appt_line_html(r, counts_by_client))
            shown += 1

    return "".join(chunks), shown


def _calendar_appt_line_html(
    row: dict[str, Any],
    counts_by_client: dict[str, int],
    *,
    long_name_from: int = 16,
) -> str:
    """HTML de una línea en celda del calendario: rejilla hora | nombre | total (compacto)."""
    hm = _appt_time_hm(row.get("appointment_date", row.get("date")))
    nm = str(row.get("customer_name") or row.get("name") or "").strip() or "—"
    short = _calendar_cell_customer_label(nm, long_from_len=long_name_from)
    cls = _client_pill_class(row, counts_by_client)
    st_cl = str(row.get("status") or "").strip().lower()
    dim_cls = " cal-appt-line--muted" if st_cl == "cancelada" else ""
    total_amt, _, _ = _financial_row_values(row)
    price_lbl = _calendar_month_value_label(total_amt)
    # Celda mensual: total en «miles» truncados, sin símbolo ni sufijo «k».
    pill_inner = (
        f'<span class="cli-pill {cls} cli-pill--cal-cell">{html_mod.escape(short)}</span>'
    )
    t_s = html_mod.escape(hm)
    staff_lbl = _assigned_staff_label(row)
    staff_part = f" · {staff_lbl}" if staff_lbl != "—" else ""
    title = html_mod.escape(
        f"{hm} · {nm}{staff_part} · Total: {_format_cop(total_amt)}"
    )
    price_esc = html_mod.escape(price_lbl)
    return (
        f"<div class='cal-appt-line{dim_cls}' title='{title}'>"
        f"<span class='cal-appt-time'>{t_s}</span>"
        f"<span class='cal-appt-pill-wrap'>{pill_inner}</span>"
        f"<span class='cal-appt-total'>{price_esc}</span>"
        "</div>"
    )


_MONTHS_ES = (
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)


def _weekday_headers_es() -> List[str]:
    return ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]


def _shift_years(base: date, years: int) -> date:
    target_year = base.year + years
    try:
        return base.replace(year=target_year)
    except ValueError:
        return base.replace(year=target_year, day=28)


def _date_range_100y_window() -> tuple[date, date]:
    today = date.today()
    return _shift_years(today, -100), _shift_years(today, 100)


def _init_appt_form_state_once() -> None:
    if st.session_state.get("_ap_form_ready"):
        return
    slot_opts = _time_slot_options()
    default_slot = "09:00" if "09:00" in slot_opts else slot_opts[0]
    defaults: Dict[str, Any] = {
        "ap_fn": "",
        "ap_ln": "",
        "ap_phone": "",
        "ap_email": "",
        "ap_ad": date.today(),
        "ap_slot": default_slot,
        "ap_det": "",
        "ap_design": "",
        "ap_dep": float(_MIN_APPOINTMENT_TOTAL_COP),
        "ap_total": float(_MIN_APPOINTMENT_TOTAL_COP),
        "ap_priority": False,
        "ap_duration_slots": 1,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    if "ap_work_kind" not in st.session_state:
        st.session_state["ap_work_kind"] = "piercing"
    if st.session_state.get("ap_work_kind") == "limpieza_tatuaje":
        st.session_state["ap_work_kind"] = "limpieza_piercing"
    if "ap_doc_type" not in st.session_state:
        st.session_state["ap_doc_type"] = "CC"
    st.session_state["_ap_form_ready"] = True


def _pop_booking_document_session() -> None:
    for k in (
        "_ap_booking_customer_id",
        "_ap_booking_customer_snapshot",
        "_ap_need_new_customer",
        "_ap_doc_verified",
        "_ap_verify_msg",
        "_ap_verify_level",
        "_ap_verified_doc_number",
        "_ap_pending_doc_type_sync",
        "ap_doc_number",
    ):
        st.session_state.pop(k, None)


def _reset_appointment_form_state() -> None:
    for key in (
        "ap_fn",
        "ap_ln",
        "ap_phone",
        "ap_email",
        "ap_ad",
        "ap_slot",
        "ap_det",
        "ap_design",
        "ap_dep",
        "ap_total",
        "ap_priority",
        "ap_work_kind",
        "ap_doc_type",
        "ap_assigned_staff_id",
        "ap_duration_slots",
        "ex_full_name",
    ):
        st.session_state.pop(key, None)
    st.session_state["_ap_form_ready"] = False
    _pop_booking_document_session()


def _initial_receipt_success_message(_dep_created: float, _service_str: str) -> str:
    """Notificación tras crear cita desde el panel (solo confirmación)."""
    return "La cita ha sido agendada."


@st.dialog("Agendar cita", width="large", dismissible=False)
def _dialog_agendar_cita() -> None:
    _init_appt_form_state_once()
    if float(st.session_state.get("ap_total") or 0) < float(_MIN_APPOINTMENT_TOTAL_COP):
        st.session_state["ap_total"] = float(_MIN_APPOINTMENT_TOTAL_COP)
    if float(st.session_state.get("ap_dep") or 0) < float(_MIN_APPOINTMENT_TOTAL_COP):
        st.session_state["ap_dep"] = float(_MIN_APPOINTMENT_TOTAL_COP)
    # Tras Verificar cliente: el tipo llega desde la API en el rerun siguiente (no escribir ap_doc_type
    # después de crear el widget con esa key — StreamlitAPIException).
    pending_doc_ty = st.session_state.pop("_ap_pending_doc_type_sync", None)
    if pending_doc_ty in ("CC", "TI", "CE", "PAS"):
        st.session_state["ap_doc_type"] = pending_doc_ty

    picked_raw = st.session_state.get("ap_ad")
    if picked_raw is None:
        st.error("Selecciona un día en el calendario para agendar.")
        if st.button("Cerrar", use_container_width=True, key="btn_appt_close_no_day"):
            st.session_state.pop("_ap_dlg", None)
            st.rerun()
        return
    picked = picked_raw if isinstance(picked_raw, date) else _parse_date(picked_raw)
    today_d = date.today()
    if picked < today_d:
        st.error("No se pueden agendar citas en fechas pasadas. Elige un día de hoy en adelante en el calendario.")
        if st.button("Cerrar", use_container_width=True, key="btn_appt_close_past_date"):
            st.session_state.pop("_ap_dlg", None)
            st.rerun()
        return

    st.markdown(
        """
        <style>
          .dlg-appt-req-banner {
            border-left: 4px solid #FF007F;
            padding: 0.5rem 0.85rem;
            margin: 0.75rem 0 0.75rem 0;
            background: rgba(255, 0, 127, 0.12);
            border-radius: 8px;
            font-size: 0.95rem;
            line-height: 1.45;
            color: #f3f4f6;
          }
          .dlg-appt-col-h {
            font-size: 0.78rem;
            font-weight: 700;
            letter-spacing: 0.04em;
            text-transform: uppercase;
            color: #A79AFF;
            margin: 0 0 0.5rem 0;
          }
        </style>
        <div class="dlg-appt-req-banner">Campos obligatorios</div>
        """,
        unsafe_allow_html=True,
    )

    c_wk, c_art = st.columns(2)
    with c_wk:
        st.markdown('<p class="dlg-appt-col-h">Tipo de trabajo</p>', unsafe_allow_html=True)
        st.selectbox(
            "¿Qué se va a realizar? *",
            options=list(_BOOKING_WORK_KIND_ORDER),
            key="ap_work_kind",
            format_func=lambda k: str(_BOOKING_WORK_KIND_META[k]["label"]),
            help="Define el servicio y qué profesional se listará (tatuador o perforador).",
        )
    wk_sel = str(st.session_state.get("ap_work_kind") or "piercing")
    if wk_sel not in _BOOKING_WORK_KIND_META:
        wk_sel = "piercing"
    need_role = _work_kind_to_assignee_role(wk_sel)
    staff_opts = [s for s in _ensure_assignable_staff() if str(s.get("role")) == need_role]

    with c_art:
        st.markdown('<p class="dlg-appt-col-h">Profesional asignado</p>', unsafe_allow_html=True)
        from streamlit_app.panel_auth import panel_auth_enabled

        assigned_id: Optional[int] = None
        role_me = str(st.session_state.get("_panel_user_role") or "")
        uid_me = st.session_state.get("_panel_user_id")
        locked_self = (
            panel_auth_enabled()
            and not st.session_state.get("_panel_session_full_access")
            and role_me == need_role
            and uid_me is not None
        )
        if locked_self:
            assigned_id = int(uid_me)
            st.session_state["ap_assigned_staff_id"] = assigned_id
            st.caption(
                "Las franjas horarias se calculan con tu disponibilidad; la cita quedará asignada a **tu usuario** del panel."
            )
        elif not staff_opts:
            st.error(
                f"No hay ningún usuario activo con rol **{need_role}** en el panel. "
                "Da de alta al profesional en **Gestión de usuarios** antes de agendar."
            )
        else:
            labels_p = [
                f"{s.get('first_name', '')} {s.get('last_name', '')} (@{s.get('username', '')})"
                for s in staff_opts
            ]
            pick_key = "ap_assigned_staff_pick"
            if pick_key not in st.session_state or st.session_state[pick_key] not in labels_p:
                st.session_state[pick_key] = labels_p[0]
            choice_p = st.selectbox(
                "Artista / profesional *",
                options=labels_p,
                key=pick_key,
                help="Cada profesional tiene su propia ocupación por día; elige quién atenderá.",
            )
            idx_p = labels_p.index(choice_p)
            assigned_id = int(staff_opts[idx_p]["id"])
            st.session_state["ap_assigned_staff_id"] = assigned_id

    c_dur, c_hr = st.columns(2)
    with c_dur:
        st.markdown('<p class="dlg-appt-col-h">Duración en agenda</p>', unsafe_allow_html=True)
        st.number_input(
            "Franjas de 30 min a reservar *",
            min_value=_MIN_BOOKING_DURATION_SLOTS,
            max_value=_MAX_BOOKING_DURATION_SLOTS,
            step=1,
            key="ap_duration_slots",
            help="Desde la hora de inicio se bloquean tantas franjas de media hora. No está ligada al tipo de trabajo.",
        )

    slot_opts = _time_slot_options()
    wk = str(st.session_state.get("ap_work_kind") or "piercing")
    if wk not in _BOOKING_WORK_KIND_META:
        wk = "piercing"
    need_slots = _booking_duration_slots_from_session()
    sched_kind = _work_kind_to_schedule_kind(wk)
    raw_appt_list = list(st.session_state.get("_ap_list") or [])
    aid_raw = st.session_state.get("ap_assigned_staff_id")
    artist_for_busy: Optional[int] = None
    if aid_raw not in (None, "", 0):
        try:
            artist_for_busy = int(aid_raw)
        except (TypeError, ValueError):
            artist_for_busy = None
    if artist_for_busy is not None:
        day_rows_cal = _appointments_for_artist_schedule(
            raw_appt_list, picked, artist_for_busy, schedule_kind=sched_kind
        )
    else:
        day_rows_cal = _appointments_same_day_schedule_kind(
            raw_appt_list, picked, sched_kind
        )
    busy_idx = _busy_slot_indices_for_day(day_rows_cal, slot_opts)
    avail_slots = _available_start_slots(slot_opts, need_slots, busy_idx)
    cur_slot = st.session_state.get("ap_slot")
    if avail_slots and cur_slot not in avail_slots:
        st.session_state["ap_slot"] = avail_slots[0]

    with c_hr:
        st.markdown('<p class="dlg-appt-col-h">Hora de inicio</p>', unsafe_allow_html=True)
        if not avail_slots:
            st.warning(
                "No quedan franjas libres ese día para esta duración. Prueba otro día o revisa las citas ya cargadas."
            )
            slot = None
        else:
            slot = st.selectbox(
                "Franja de inicio *",
                options=avail_slots,
                key="ap_slot",
                help=f"Se reservan {need_slots} franja(s) de 30 min desde esta hora.",
            )
            st.caption(f"Inicio **{slot}** hora local (duración **{need_slots * 30}** min).")

    st.markdown(
        f"**Fecha de la cita:** {picked.strftime('%d/%m/%Y')} _(elegida en el calendario)_"
    )

    st.markdown('<p class="dlg-appt-col-h">Verificación de documento</p>', unsafe_allow_html=True)
    c_doc_l, c_doc_r = st.columns(2)
    with c_doc_l:
        st.selectbox(
            "Tipo de documento *",
            options=["CC", "TI", "CE", "PAS"],
            format_func=lambda x: {
                "CC": "CC — Cédula",
                "TI": "TI — Tarjeta de identidad",
                "CE": "CE — Extranjería",
                "PAS": "PAS — Pasaporte",
            }[x],
            key="ap_doc_type",
            help="Para clientes nuevos y como tipo de documento al verificar o registrar.",
        )
        st.text_input(
            "Número de documento *",
            key="ap_doc_number",
            placeholder="Sin puntos ni espacios, si es posible",
        )
    with c_doc_r:
        st.markdown("<div style='height:4.5rem'></div>", unsafe_allow_html=True)
        if st.button("Verificar documento", type="secondary", use_container_width=True, key="ap_btn_verify_doc"):
            doc_in = (st.session_state.get("ap_doc_number") or "").strip()
            if len(doc_in) < 5:
                st.session_state["_ap_verify_level"] = "error"
                st.session_state["_ap_verify_msg"] = "Ingresa un documento válido (mínimo 5 caracteres)."
                st.session_state["_ap_doc_verified"] = False
            else:
                ok_f, msg_f, row_f = fetch_customer_by_document(doc_in)
                if not ok_f:
                    st.session_state["_ap_verify_level"] = "error"
                    st.session_state["_ap_verify_msg"] = msg_f
                    st.session_state["_ap_doc_verified"] = False
                elif msg_f == "not_found":
                    st.session_state["_ap_booking_customer_id"] = None
                    st.session_state["_ap_need_new_customer"] = True
                    st.session_state["_ap_doc_verified"] = True
                    st.session_state["_ap_verified_doc_number"] = doc_in
                    st.session_state["_ap_verify_level"] = "warning"
                    st.session_state["_ap_verify_msg"] = (
                        "Cliente no registrado. Completa nombre, apellido, celular y correo. "
                        "La fecha de nacimiento y el tutor (si aplica) se registran al firmar el contrato o en la ficha del cliente."
                    )
                else:
                    st.session_state["_ap_booking_customer_id"] = int(row_f["id"])
                    st.session_state["_ap_need_new_customer"] = False
                    st.session_state["_ap_doc_verified"] = True
                    st.session_state["_ap_verified_doc_number"] = doc_in
                    st.session_state["_ap_booking_customer_snapshot"] = dict(row_f)
                    st.session_state["ap_fn"] = str(row_f.get("first_name") or "")
                    st.session_state["ap_ln"] = str(row_f.get("last_name") or "")
                    st.session_state["ap_phone"] = str(row_f.get("phone_number") or "")
                    st.session_state["ap_email"] = str(row_f.get("email") or "")
                    raw_dt = str(row_f.get("document_type") or "").strip().upper()
                    if raw_dt in ("CC", "TI", "CE", "PAS"):
                        st.session_state["_ap_pending_doc_type_sync"] = raw_dt
                    st.session_state["_ap_verify_level"] = "success"
                    st.session_state["_ap_verify_msg"] = f"Cliente encontrado (id {row_f['id']}). Datos cargados."
            st.rerun()

    v_lvl = st.session_state.get("_ap_verify_level")
    v_msg = st.session_state.get("_ap_verify_msg")
    if v_msg and v_lvl:
        if v_lvl == "error":
            st.error(v_msg)
        elif v_lvl == "success":
            st.success(v_msg)
        else:
            st.warning(v_msg)

    if st.session_state.get("_ap_need_new_customer"):
        st.caption(
            "**Tarjeta de identidad (TI)** u otros documentos: se admite al agendar. "
            "La fecha de nacimiento y el estado de menor/tutor se definen al completar la ficha o en la firma del contrato."
        )

    st.markdown('<p class="dlg-appt-col-h">Cliente</p>', unsafe_allow_html=True)
    cl1, cl2 = st.columns(2)
    with cl1:
        fn = st.text_input("Nombre *", key="ap_fn")
        phone = st.text_input(
            "Celular *",
            key="ap_phone",
            help="10 dígitos; puedes incluir espacios o prefijo, se cuentan solo los números.",
        )
    with cl2:
        ln = st.text_input("Apellido *", key="ap_ln")
        st.text_input("Correo electrónico *", key="ap_email")

    st.markdown('<p class="dlg-appt-col-h">Cita y montos</p>', unsafe_allow_html=True)
    cm1, cm2 = st.columns(2)
    with cm1:
        st.text_area(
            "Descripción del diseño (opcional)",
            height=68,
            key="ap_design",
            help="Se guarda en el detalle de la cita junto con las observaciones.",
        )
        st.text_area(
            "Notas u observaciones (opcional)",
            height=68,
            key="ap_det",
            help="Texto adicional (indicaciones, zona, etc.).",
        )
        st.checkbox(
            "Cita prioritaria",
            key="ap_priority",
            help="Se muestra con etiqueta roja en calendario y listado (prevalece sobre cliente nuevo/recurrente salvo reprogramación).",
        )
    with cm2:
        total_amount = st.number_input(
            "Valor total del trabajo (COP) *",
            min_value=float(_MIN_APPOINTMENT_TOTAL_COP),
            step=5000.0,
            format="%.0f",
            key="ap_total",
        )
        deposit = st.number_input(
            "Saldo abonado (COP) *",
            min_value=float(_MIN_APPOINTMENT_TOTAL_COP),
            step=5000.0,
            format="%.0f",
            key="ap_dep",
            help=f"Mínimo {_format_cop(_MIN_APPOINTMENT_TOTAL_COP)}.",
        )
        pending_balance = round(float(total_amount) - float(deposit), 2)
        st.caption(f"Saldo pendiente calculado: {_format_cop(max(pending_balance, 0))}")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Crear cita", type="primary", use_container_width=True, key="btn_appt_create"):
            if not st.session_state.get("_ap_doc_verified"):
                st.error("Debes verificar el documento antes de crear la cita.")
                return
            doc_in = (st.session_state.get("ap_doc_number") or "").strip()
            if len(doc_in) < 5:
                st.error("El número de documento no es válido.")
                return
            snap = (st.session_state.get("_ap_verified_doc_number") or "").strip()
            if snap and snap != doc_in:
                st.error("El documento cambió respecto al verificado. Pulsa de nuevo «Verificar documento».")
                return
            cust_id = st.session_state.get("_ap_booking_customer_id")
            need_new = bool(st.session_state.get("_ap_need_new_customer"))
            aid_submit = st.session_state.get("ap_assigned_staff_id")
            if aid_submit is None or aid_submit == "":
                st.error("Indica el **profesional** que atenderá la cita (tatuador o perforador).")
                return
            aid_int = int(aid_submit)
            wk_submit = str(st.session_state.get("ap_work_kind") or "piercing")
            if wk_submit not in _BOOKING_WORK_KIND_META:
                wk_submit = "piercing"
            need_slots_submit = _booking_duration_slots_from_session()
            sched_submit = _work_kind_to_schedule_kind(wk_submit)
            slot_opts_chk = _time_slot_options()
            raw_chk = list(st.session_state.get("_ap_list") or [])
            day_chk = _appointments_for_artist_schedule(
                raw_chk, picked, aid_int, schedule_kind=sched_submit
            )
            busy_chk = _busy_slot_indices_for_day(day_chk, slot_opts_chk)
            avail_chk = _available_start_slots(slot_opts_chk, need_slots_submit, busy_chk)
            if not avail_chk:
                st.error("No hay franja disponible ese día para la duración de este trabajo.")
                return
            slot_str = (st.session_state.get("ap_slot") or "").strip()
            if slot_str not in avail_chk:
                st.error("La franja elegida ya no está libre. Vuelve a seleccionar la hora.")
                return
            detail_raw = _booking_observations_and_design_for_api()
            total_amount = float(int(round(float(st.session_state.get("ap_total") or 0))))
            deposit = float(int(round(float(st.session_state.get("ap_dep") or 0))))
            service, detail_for_api = _service_and_detail_for_work_kind(wk_submit, detail_raw)
            detail_for_api = _append_agenda_slots_marker(detail_for_api, need_slots_submit)
            full_name = f"{(fn or '').strip()} {(ln or '').strip()}".strip()
            dt_str = _combine_appointment_datetime(picked, slot_str)
            email_s = (st.session_state.get("ap_email") or "").strip()
            valid, errs = validate_appointment(
                full_name,
                (phone or "").strip(),
                email_s,
                service,
                dt_str,
                detail_raw,
                deposit,
            )
            if not valid:
                _show_validation_errors(errs)
                return
            if deposit > total_amount:
                st.error("El saldo abonado no puede ser mayor que el valor total del trabajo.")
                return
            if deposit < float(_MIN_APPOINTMENT_TOTAL_COP):
                st.error(
                    f"El saldo abonado debe ser al menos {_format_cop(_MIN_APPOINTMENT_TOTAL_COP)}."
                )
                return
            if picked < today_d:
                st.error("La fecha de la cita no puede ser anterior a hoy.")
                return

            dep_norm = max(0.0, float(int(round(float(deposit)))))
            total_int = float(int(round(float(total_amount))))
            if total_int < float(_MIN_APPOINTMENT_TOTAL_COP):
                st.error(
                    f"El valor total del trabajo debe ser al menos {_format_cop(_MIN_APPOINTMENT_TOTAL_COP)}."
                )
                return
            appt_payload: Dict[str, Any] = {
                "name": full_name,
                "phone": (phone or "").strip(),
                "service": (service or "").strip(),
                "date": dt_str,
                "detail": detail_for_api,
                "deposit": dep_norm,
                "total_amount": total_int,
                "pending_balance": float(max(round(total_int - dep_norm, 2), 0)),
                "is_priority": bool(st.session_state.get("ap_priority")),
                "assigned_panel_user_id": aid_int,
            }
            if cust_id is not None:
                appt_payload["customer_id"] = int(cust_id)
                snap = st.session_state.get("_ap_booking_customer_snapshot")
                if not isinstance(snap, dict) or int(snap.get("id") or 0) != int(cust_id):
                    st.error(
                        "Los datos del cliente no coinciden con la verificación. "
                        "Pulsa **Verificar documento** de nuevo."
                    )
                    return
                try:
                    c_exist = _booking_customer_create_for_existing_client(
                        snap,
                        first_name=(fn or "").strip(),
                        last_name=(ln or "").strip(),
                        phone_number=(phone or "").strip(),
                        email_s=email_s,
                        document_number=doc_in,
                    )
                except ValidationError as ve:
                    st.error(str(ve))
                    return
                appt_payload["customer"] = c_exist.model_dump(mode="json")
            elif need_new:
                doc_ty = str(st.session_state.get("ap_doc_type") or "CC")
                if doc_ty not in ("CC", "TI", "CE", "PAS"):
                    doc_ty = "CC"
                try:
                    c_new = CustomerCreate(
                        first_name=(fn or "").strip(),
                        last_name=(ln or "").strip(),
                        birth_date=CUSTOMER_BIRTH_PENDING,
                        document_type=doc_ty,  # type: ignore[arg-type]
                        document_number=doc_in,
                        document_issue_date=None,
                        email=email_s,
                        phone_number=(phone or "").strip(),
                        address=None,
                        is_minor=False,
                        guardian_name=None,
                        guardian_document_type=None,
                        guardian_document_number=None,
                        guardian_document_issue_date=None,
                    )
                except ValidationError as ve:
                    st.error(str(ve))
                    return
                appt_payload["customer"] = c_new.model_dump(mode="json")
            else:
                st.error("Verifica el documento antes de crear la cita.")
                return

            with st.spinner("Guardando cita…"):
                ok_a, code_a, data_a = api_client.post_appointment(appt_payload)
            if ok_a:
                st.session_state["_ap_reload"] = True
                dep_created = max(0.0, round(float(appt_payload.get("deposit") or 0), 2))
                ok_msg = _initial_receipt_success_message(dep_created, str(appt_payload.get("service") or ""))
                _queue_appointment_action_success(ok_msg)
                _reset_appointment_form_state()
                st.session_state.pop("_ap_dlg", None)
                st.rerun()
            else:
                st.toast(
                    f"Error HTTP {code_a}: {_api_error(data_a)}",
                    icon="❌",
                    duration="long",
                )
    with c2:
        if st.button("Cancelar", use_container_width=True, key="btn_appt_cancel"):
            _reset_appointment_form_state()
            st.session_state.pop("_ap_dlg", None)
            st.rerun()


def _calendar_fragment_rerun() -> None:
    """Reejecutar solo el fragment del calendario si la versión de Streamlit lo permite."""
    try:
        st.rerun(scope="fragment")
    except TypeError:
        st.rerun()


@st.fragment
def _render_main_calendar(
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
    counts_by_client: dict[str, int],
    *,
    team_layout: bool = False,
) -> None:
    """Vista principal: mes con citas por día; al pulsar el día se abre el diálogo de agendamiento."""
    ym_key = "_ap_cal_ym"
    today = date.today()
    if ym_key not in st.session_state:
        st.session_state[ym_key] = (today.year, today.month)
    y, m = st.session_state[ym_key]

    st.markdown("##### Calendario de citas")

    n1, n2, n3 = st.columns([1, 3, 1])
    with n1:
        if st.button("◀ Mes", key="cal_main_prev_m"):
            _clear_calendar_dialog_focus()
            if m <= 1:
                st.session_state[ym_key] = (y - 1, 12)
            else:
                st.session_state[ym_key] = (y, m - 1)
            _calendar_fragment_rerun()
    with n2:
        st.markdown(
            f"<div style='text-align:center;font-weight:600;font-size:1.05rem'>{_MONTHS_ES[m]} {y}</div>",
            unsafe_allow_html=True,
        )
    with n3:
        if st.button("Mes ▶", key="cal_main_next_m"):
            _clear_calendar_dialog_focus()
            if m >= 12:
                st.session_state[ym_key] = (y + 1, 1)
            else:
                st.session_state[ym_key] = (y, m + 1)
            _calendar_fragment_rerun()

    hdr_cells = st.columns(7, gap=None)
    for i, lab in enumerate(_weekday_headers_es()):
        hdr_cells[i].markdown(
            f"<div class='cal-m-whcell'>{html_mod.escape(lab)}</div>",
            unsafe_allow_html=True,
        )
    weeks_grid = calendar.monthcalendar(y, m)
    allow_footer_book = _panel_is_technician_role() is False
    for week in weeks_grid:
        row_cells = st.columns(7, gap=None)
        for i, d in enumerate(week):
            with row_cells[i]:
                if d == 0:
                    st.markdown(
                        "<div class='cal-cell-spacer cal-cell-spacer--month'></div>",
                        unsafe_allow_html=True,
                    )
                else:
                    day_rows = buckets.get((y, m, d), [])
                    picked_cell = date(y, m, d)
                    today_cls = " cal-cell-today" if picked_cell == today else ""
                    body, _shown = _calendar_day_inner_body_html(
                        day_rows,
                        counts_by_client,
                        team_layout=team_layout,
                    )
                    is_past = picked_cell < date.today()
                    footer_off = is_past or not allow_footer_book
                    fh = html_mod.escape(
                        "Los tatuadores y perforadores no pueden agendar desde el calendario"
                        if not allow_footer_book
                        else (
                            "No se pueden agendar citas en fechas pasadas"
                            if is_past
                            else f"Agendar cita · {picked_cell.strftime('%d/%m/%Y')}"
                        )
                    )
                    footer_html = _calendar_month_footer_strip_html(
                        y,
                        m,
                        d,
                        disabled_footer=footer_off,
                        title_esc=fh,
                    )
                    st.markdown(
                        f"<div class='cal-cell cal-cell--month{today_cls}'>"
                        f"<div class='cal-cell-head'><span class='cal-cell-daynum'>"
                        f"{html_mod.escape(str(d))}</span></div>"
                        f"<div class='cal-day-inner'>{body}</div>"
                        f"{footer_html}</div>",
                        unsafe_allow_html=True,
                    )
    _render_calendar_month_hidden_book_widgets(y, m, weeks_grid)
    _inject_calendar_month_footer_bridge()


def _calendar_overflow_day_sheet_link_row(
    r: dict[str, Any],
    hist_counts: dict[str, int],
    *,
    key_suffix: str,
) -> None:
    """Resumen rápido + acceso al panel único de ficha/detalle."""
    st.markdown(_calendar_overflow_row_html(r, hist_counts), unsafe_allow_html=True)
    aid = int(r.get("id", 0) or 0)
    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button("Ficha completa", use_container_width=True, key=f"cal_day_open_sheet_{key_suffix}"):
            st.session_state["_cal_focus_appt_id"] = aid
            st.session_state.pop("_cal_overflow_day", None)
            st.rerun()


def _render_calendar_focus_appointment_body(r: dict[str, Any], hist_counts: dict[str, int]) -> None:
    """Formulario único para ver/editar cita desde calendario (rejilla día / semana)."""
    aid = int(r.get("id", 0) or 0)
    if aid <= 0:
        return
    status_l = str(r.get("status") or "Agendada").strip()
    montos_locked = aid <= 0 or status_l not in {"Agendada", "Reprogramada"}
    firmar_disabled = _firmar_contrato_disabled(r)
    firma_lbl = _firmar_contrato_button_label(r)
    tech = _panel_is_technician_role()
    aid_s = str(aid)

    if tech:
        st.markdown(_calendar_overflow_row_html(r, hist_counts), unsafe_allow_html=True)
        if st.button(firma_lbl, use_container_width=True, key=f"fcd_tech_{aid_s}_firma", disabled=firmar_disabled):
            _clear_calendar_dialog_focus()
            _open_firma_contrato_nav(r, aid)
        return

    seed_k = "_fcd_seed_appt_id"
    slot_opts = _time_slot_options()
    today_d = date.today()
    if st.session_state.get(seed_k) != aid:
        st.session_state[seed_k] = aid
        _, sl0 = _parse_existing_slot(r.get("appointment_date", r.get("date")))
        if sl0 not in slot_opts and slot_opts:
            sl0 = slot_opts[0]
        dur0 = _duration_slots_for_existing_appointment(r)
        last0 = _appointment_last_start_slot(sl0, dur0, slot_opts)
        if last0 not in slot_opts:
            last0 = sl0
        plain = _appointment_detail_plain_body(str(r.get("detail") or ""))
        dz0, obs0 = _split_design_obs_plain(plain)
        st.session_state[f"fcd_start_{aid_s}"] = sl0
        st.session_state[f"fcd_last_{aid_s}"] = last0
        st.session_state[f"fcd_pri_{aid_s}"] = bool(_row_is_priority(r))
        st.session_state[f"fcd_tot_{aid_s}"] = max(
            int(_MIN_APPOINTMENT_TOTAL_COP), int(round(float(r.get("total_amount") or 0)))
        )
        st.session_state[f"fcd_dz_{aid_s}"] = dz0
        st.session_state[f"fcd_obs_{aid_s}"] = obs0
        wk_here = _work_kind_infer_from_existing_row(r)
        role_need = _work_kind_to_assignee_role(wk_here)
        staff_opts = [s for s in _ensure_assignable_staff() if str(s.get("role")) == role_need]
        if not staff_opts:
            staff_opts = list(_ensure_assignable_staff())
        lbls_o: dict[int, str] = {}
        opt_ids_o: list[int] = []
        for s in staff_opts:
            try:
                oid = int(s.get("id") or 0)
            except (TypeError, ValueError):
                continue
            if oid <= 0:
                continue
            opt_ids_o.append(oid)
            fn = str(s.get("first_name") or "").strip()
            ln = str(s.get("last_name") or "").strip()
            lab = (fn + " " + ln).strip() or str(s.get("username") or f"#{oid}")
            lbls_o[oid] = lab + f" (@{s.get('username') or oid})"
        cur_id: Optional[int] = None
        raw_as = r.get("assigned_panel_user_id")
        if raw_as not in (None, "", 0):
            try:
                cur_id = int(raw_as)
            except (TypeError, ValueError):
                cur_id = None
        if cur_id is not None and cur_id > 0 and cur_id not in lbls_o:
            lbls_o[cur_id] = _assigned_artist_display_name(r) + " (historial)"
            opt_ids_o = [cur_id] + [x for x in opt_ids_o if x != cur_id]
        pick_art = cur_id if cur_id is not None and cur_id in lbls_o else (opt_ids_o[0] if opt_ids_o else None)
        st.session_state[f"fcd_artpick_{aid_s}"] = int(pick_art) if pick_art is not None else 0

        day_appt, hm_orig = _parse_existing_slot(r.get("appointment_date", r.get("date")))
        hm_eff = hm_orig if hm_orig != "—" else "09:00"
        dt_orig = _combine_appointment_datetime(day_appt, hm_eff)
        base = {
            "start": sl0,
            "last": last0,
            "prio": bool(_row_is_priority(r)),
            "tot": max(int(_MIN_APPOINTMENT_TOTAL_COP), int(round(float(r.get("total_amount") or 0)))),
            "artist": pick_art,
            "dz": dz0,
            "obs": obs0,
            "dt_orig": dt_orig,
            "dur": dur0,
        }
        st.session_state[f"fcd_base_{aid_s}"] = base

    hdr_l, hdr_r = st.columns([3, 2])
    with hdr_l:
        st.markdown("### Cita")
    with hdr_r:
        st.markdown(
            f"<div style='text-align:right'>{_format_appt_created_at(r.get('created_at'))}<br>"
            f"{_status_pill_html(status_l)}</div>",
            unsafe_allow_html=True,
        )

    cust: Optional[dict[str, Any]] = None
    try:
        cid_raw = r.get("customer_id")
        if cid_raw:
            cid = int(cid_raw)
            ok_c, _, cdata = api_client.get_customer(cid)
            if ok_c and isinstance(cdata, dict):
                cust = cdata
    except (TypeError, ValueError):
        cust = None

    if cust:
        nm = (
            str(cust.get("first_name") or "").strip() + " " + str(cust.get("last_name") or "").strip()
        ).strip() or str(r.get("customer_name") or "").strip()
        doc_lbl = str(cust.get("document_type") or "CC")
        doc_n = str(cust.get("document_number") or "—").strip()
        em = str(cust.get("email") or "—").strip()
        cel = str(cust.get("phone_number") or r.get("phone") or "—").strip()
        st.markdown(f"**Cliente:** {html_mod.escape(nm or '—')}")
        st.caption(f"{doc_lbl} {doc_n} · ✉ {em} · 📱 {cel}")
    else:
        nm2 = str(r.get("customer_name") or r.get("name") or "—").strip()
        cel2 = str(r.get("phone") or "—").strip()
        st.markdown(f"**Cliente:** {html_mod.escape(nm2)} _(sin cliente vinculado o no cargado)_")
        st.caption(f"📱 Teléfono en cita: **{cel2}**")

    st.divider()

    repro_lock = _reprogram_disabled_for_row(r) or status_l in {"Cancelada", "Finalizada"}
    t1a, t1b = st.columns(2)
    with t1a:
        st.selectbox(
            "Desde (inicio)",
            options=slot_opts,
            disabled=repro_lock,
            key=f"fcd_start_{aid_s}",
        )
    with t1b:
        st.selectbox(
            "Hasta (última franja ocupada)",
            options=slot_opts,
            disabled=repro_lock,
            key=f"fcd_last_{aid_s}",
        )
    if repro_lock:
        st.caption("Horario bloqueado: estado cerrado o la cita no admite cambio de franja desde aquí.")

    st.divider()

    srv_l, srv_r = st.columns([4, 2])
    with srv_l:
        st.markdown("### Cita · servicio")
    rc_key = f"{_AP_RECEIPTS_CACHE_PREFIX}{aid}"
    cached_r = st.session_state.get(rc_key)
    if not isinstance(cached_r, tuple) or len(cached_r) != 3:
        cached_r = api_client.get_appointment_receipts(aid)
        st.session_state[rc_key] = cached_r
    ok_rr, _, rrows = cached_r
    rec_txt = "—"
    if ok_rr and isinstance(rrows, list):
        ids_r = sorted(
            int(x.get("id") or 0) for x in rrows if isinstance(x, dict) and int(x.get("id") or 0)
        )
        if ids_r:
            rec_txt = ", ".join(str(x) for x in ids_r)
    with srv_r:
        st.caption(f"**Recibo id(s)** — solo lectura: {rec_txt}")

    artist_locked = bool(r.get("has_signed_contract"))
    lbls_reload = {}
    role_need_r = _work_kind_to_assignee_role(_work_kind_infer_from_existing_row(r))
    staff_reload = [s for s in _ensure_assignable_staff() if str(s.get("role")) == role_need_r] or list(
        _ensure_assignable_staff()
    )
    opt_ids_reload: list[int] = []
    for s in staff_reload:
        try:
            oid = int(s.get("id") or 0)
        except (TypeError, ValueError):
            continue
        if oid <= 0:
            continue
        fn = str(s.get("first_name") or "").strip()
        ln = str(s.get("last_name") or "").strip()
        lab = ((fn + " " + ln).strip() or str(s.get("username") or oid)) + f" (@{s.get('username') or oid})"
        lbls_reload[oid] = lab
        opt_ids_reload.append(oid)

    sel_art = int(st.session_state.get(f"fcd_artpick_{aid_s}") or 0)
    if sel_art not in lbls_reload and sel_art > 0:
        lbls_reload[sel_art] = _assigned_artist_display_name(r)
        opt_ids_reload = [sel_art] + [x for x in opt_ids_reload if x != sel_art]
    ix_def = (
        opt_ids_reload.index(sel_art)
        if sel_art in opt_ids_reload
        else 0
    )
    if not opt_ids_reload:
        st.warning("No hay artistas configurados en el panel para este tipo de servicio.")
    else:
        st.selectbox(
            "Artista",
            options=opt_ids_reload,
            index=min(ix_def, max(0, len(opt_ids_reload) - 1)),
            format_func=lambda i: lbls_reload.get(int(i), str(i)),
            disabled=montos_locked or artist_locked,
            key=f"fcd_artpick_{aid_s}",
        )

    tot_min = max(int(_MIN_APPOINTMENT_TOTAL_COP), 1)
    st.number_input(
        "Valor del tatuaje / trabajo (COP, entero)",
        min_value=float(tot_min),
        step=10000.0,
        disabled=montos_locked,
        key=f"fcd_tot_{aid_s}",
        format="%.0f",
    )

    _ = st.checkbox("Cita prioritaria", disabled=montos_locked, key=f"fcd_pri_{aid_s}")

    _ = st.text_area("Descripción del diseño", disabled=montos_locked, height=90, key=f"fcd_dz_{aid_s}")
    _ = st.text_area("Observaciones", disabled=montos_locked, height=80, key=f"fcd_obs_{aid_s}")

    if artist_locked:
        st.caption("El artista no se puede cambiar si ya existe contrato firmado en esta cita.")

    st.divider()
    st.markdown("##### Abonos")
    okp, pcode, pays_raw = _get_appointment_payments_cached(aid)
    if okp and isinstance(pays_raw, list) and pays_raw:
        rows_view: list[dict[str, str]] = []
        for pr in pays_raw:
            pid = int(pr.get("id") or 0)
            amt = float(pr.get("amount") or 0)
            pon = pr.get("paid_on")
            if hasattr(pon, "isoformat"):
                d_show = pon.isoformat()[:10]  # type: ignore[attr-defined]
            elif isinstance(pon, str) and len(pon) >= 10:
                d_show = pon[:10]
            else:
                ca = pr.get("created_at")
                ds = str(ca or "")[:10]
                d_show = ds if ds else "—"
            rows_view.append({"id": pid, "fecha_abono": d_show, "valor": _format_cop(amt), "nota": str(pr.get("note") or "")})
        st.dataframe(rows_view, use_container_width=True, hide_index=True)
    elif not okp:
        st.warning(f"No se cargaron abonos (HTTP {pcode}).")
    else:
        st.caption("Aún no hay filas en el historial de abonos.")

    if montos_locked is False:
        new_pd, ncol = st.columns([2, 2])
        with new_pd:
            st.date_input("Fecha en que abona", min_value=today_d, format="DD/MM/YYYY", key=f"fcd_newpay_dt_{aid_s}")
        with ncol:
            st.number_input(
                "Valor nuevo abono (COP)",
                min_value=float(_MIN_APPOINTMENT_TOTAL_COP),
                step=5000.0,
                key=f"fcd_newpay_am_{aid_s}",
                format="%.0f",
            )
        if st.button("➕ Agregar abono", key=f"fcd_newpay_btn_{aid_s}"):
            paid_date = st.session_state.get(f"fcd_newpay_dt_{aid_s}")
            amt_add = float(int(st.session_state.get(f"fcd_newpay_am_{aid_s}") or 0))
            if amt_add < float(_MIN_APPOINTMENT_TOTAL_COP):
                st.toast("El abono debe ser al menos el mínimo configurado.", icon="⚠️")
            else:
                iso = paid_date.isoformat() if hasattr(paid_date, "isoformat") else None
                with st.spinner("Registrando abono…"):
                    ok_pay, cd_pay, body_pay = api_client.post_appointment_payment(
                        aid, amt_add, note=None, paid_on=iso
                    )
                if ok_pay:
                    _purge_appointment_payment_caches()
                    st.session_state.pop(f"{_AP_FIN_PAYMENTS_CACHE_PREFIX}{aid}", None)
                    st.session_state.pop(seed_k, None)
                    st.session_state["_ap_reload"] = True
                    _queue_appointment_action_success(
                        "**Abono registrado.** Revisa PDF en **Recibos** cuando corresponda."
                    )
                    st.rerun()
                else:
                    st.toast(f"Error HTTP {cd_pay}: {_api_error(body_pay)}", icon="❌")

        pay_ids = []
        if okp and isinstance(pays_raw, list):
            pay_ids = [int(p.get("id") or 0) for p in pays_raw if int(p.get("id") or 0)]
        lbl_pay = {"—": 0}
        for pr in pays_raw if isinstance(pays_raw, list) else []:
            pid = int(pr.get("id") or 0)
            if pid <= 0:
                continue
            amt = float(pr.get("amount") or 0)
            lbl_pay[f"#{pid} · {_format_cop(amt)}"] = pid
        pick_lbl = st.selectbox("Editar un abono", options=list(lbl_pay.keys()), key=f"fcd_paypicklbl_{aid_s}")
        pid_edit = lbl_pay[pick_lbl]
        if pid_edit:
            pr_hit = next(
                (
                    z
                    for z in pays_raw
                    if isinstance(z, dict) and int(z.get("id") or 0) == int(pid_edit)
                ),
                None,
            )
            if isinstance(pr_hit, dict):
                st.number_input(
                    "Nuevo valor (COP)",
                    min_value=1.0,
                    step=1000.0,
                    key=f"fcd_patched_am_{aid_s}_{pid_edit}",
                    format="%.0f",
                )
                st.date_input("Fecha efectiva abono", min_value=today_d, format="DD/MM/YYYY", key=f"fcd_patched_dt_{aid_s}_{pid_edit}")  # type: ignore[arg-type]
                if st.button(f"Actualizar abono #{pid_edit}", key=f"fcd_patched_sv_{aid_s}_{pid_edit}"):
                    n_am = float(int(st.session_state.get(f"fcd_patched_am_{aid_s}_{pid_edit}") or 0))
                    nd_dt = st.session_state.get(f"fcd_patched_dt_{aid_s}_{pid_edit}")
                    nd_iso = nd_dt.isoformat()[:10] if hasattr(nd_dt, "isoformat") else None
                    ok_u, cu, bu = api_client.patch_appointment_payment(
                        aid,
                        pid_edit,
                        amount=n_am,
                        paid_on=nd_iso,
                    )
                    if ok_u:
                        _purge_appointment_payment_caches()
                        st.session_state.pop(f"{_AP_FIN_PAYMENTS_CACHE_PREFIX}{aid}", None)
                        st.session_state.pop(seed_k, None)
                        st.session_state["_ap_reload"] = True
                        _queue_appointment_action_success("**Abono actualizado.**")
                        st.rerun()
                    else:
                        st.toast(f"Error HTTP {cu}: {_api_error(bu)}", icon="❌")

    b_now = dict(st.session_state.get(f"fcd_base_{aid_s}") or {})
    cur_start_s = str(st.session_state.get(f"fcd_start_{aid_s}") or "")
    cur_last_s = str(st.session_state.get(f"fcd_last_{aid_s}") or "")
    cur_pri_b = bool(st.session_state.get(f"fcd_pri_{aid_s}") or False)
    cur_tot_f = float(st.session_state.get(f"fcd_tot_{aid_s}") or 0)
    cur_art_i = int(st.session_state.get(f"fcd_artpick_{aid_s}") or 0)
    dz_cur = str(st.session_state.get(f"fcd_dz_{aid_s}") or "")
    obs_cur = str(st.session_state.get(f"fcd_obs_{aid_s}") or "")
    dur_now = (
        _duration_slots_from_start_last(cur_start_s, cur_last_s, slot_opts)
        if cur_start_s in slot_opts and cur_last_s in slot_opts
        else int(b_now.get("dur") or _MIN_BOOKING_DURATION_SLOTS)
    )
    ds_day_appt, hm_row = _parse_existing_slot(r.get("appointment_date", r.get("date")))
    new_booking_dt = (
        _combine_appointment_datetime(ds_day_appt, cur_start_s) if cur_start_s != "—" else str(b_now.get("dt_orig") or "")
    )
    base_dur = int(b_now.get("dur") or _MIN_BOOKING_DURATION_SLOTS)
    sched_need = montos_locked is False and (
        (cur_start_s, cur_last_s) != (b_now.get("start"), b_now.get("last"))
        or dur_now != base_dur
        or (new_booking_dt and new_booking_dt != str(b_now.get("dt_orig") or ""))
    )
    text_dirty = dz_cur != str(b_now.get("dz") or "") or obs_cur != str(b_now.get("obs") or "")
    art_dirty = cur_art_i != int(b_now.get("artist") or 0)
    prio_dirty = cur_pri_b != bool(b_now.get("prio"))
    tot_dirty_cop = (
        montos_locked is False and int(round(cur_tot_f)) != int(b_now.get("tot") or 0)
    )
    form_dirty = bool(b_now) and (
        sched_need or text_dirty or art_dirty or prio_dirty or tot_dirty_cop
    )

    st.divider()
    st.markdown("##### Acciones")

    leave_confirm = bool(st.session_state.get(f"fcd_confirm_leave_{aid_s}"))
    if leave_confirm:
        st.warning("Hay cambios sin guardar. Si sales ahora se descartan en esta pantalla (no en el servidor).")
        cx1, cx2 = st.columns(2)
        with cx1:
            if st.button("Salir sin guardar", use_container_width=True, key=f"fcd_lvy_{aid_s}"):
                st.session_state.pop(f"fcd_confirm_leave_{aid_s}", None)
                st.session_state.pop(seed_k, None)
                _clear_calendar_dialog_focus()
                st.rerun()
        with cx2:
            if st.button("Seguir editando", use_container_width=True, key=f"fcd_lvn_{aid_s}"):
                st.session_state.pop(f"fcd_confirm_leave_{aid_s}", None)
                st.rerun()
        return

    detail_for_save = _rebuild_detail_for_patch(r, dz_cur, obs_cur, agenda_slots_override=dur_now)
    detail_meta_only = detail_for_save if (not sched_need and text_dirty) else None

    if montos_locked is False:

        def _fcd_save_all() -> None:
            if tot_dirty_cop:
                nf = _find_appointment_row_by_id(aid) or r
                dep_live = float(nf.get("deposit") or 0)
                ttot = float(int(round(cur_tot_f)))
                if dep_live > ttot + 0.01:
                    st.toast("El valor total no puede ser menor que lo ya abonado.", icon="⚠️")
                    return
                pend = round(ttot - dep_live, 2)
                ok_t, ct, bt = api_client.patch_appointment_financials(aid, ttot, dep_live, pend)
                if not ok_t:
                    st.toast(f"Montos (HTTP {ct}): {_api_error(bt)}", icon="❌")
                    return
            if sched_need:
                ok_s, cs, bs = api_client.patch_appointment_reschedule(aid, new_booking_dt, detail_for_save)
                if not ok_s:
                    st.toast(f"Horario (HTTP {cs}): {_api_error(bs)}", icon="❌")
                    return
            meta_need = art_dirty or prio_dirty or detail_meta_only is not None
            if meta_need:
                ok_m, cm, bm = api_client.patch_appointment_meta(
                    aid,
                    assigned_panel_user_id=(cur_art_i if art_dirty else None),
                    is_priority=cur_pri_b,
                    detail=detail_meta_only,
                )
                if not ok_m:
                    st.toast(f"Servicio/meta (HTTP {cm}): {_api_error(bm)}", icon="❌")
                    return
            st.session_state.pop(seed_k, None)
            st.session_state["_ap_reload"] = True
            _queue_appointment_action_success("**Cambios guardados en la cita.**")
            st.rerun()

        ba1, ba2, ba3, ba4 = st.columns(4)
        with ba1:
            if st.button("Guardar cambios", type="primary", use_container_width=True, key=f"fcd_sv_{aid_s}"):
                _fcd_save_all()
        with ba2:
            anular_disabled = aid <= 0 or status_l in {"Cancelada", "Finalizada"}
            if st.button(
                "Cancelar cita",
                use_container_width=True,
                disabled=anular_disabled,
                key=f"fcd_can_{aid_s}",
            ):
                st.session_state.pop(seed_k, None)
                st.session_state.pop("_cal_focus_appt_id", None)
                st.session_state["_ap_cancel_item"] = dict(r)
                st.rerun()
        with ba3:
            if st.button(
                "Enviar recibo",
                use_container_width=True,
                key=f"fcd_rc_{aid_s}",
            ):
                st.session_state["_ap_receipts_item"] = dict(r)
                st.session_state.pop(seed_k, None)
                st.session_state.pop("_cal_focus_appt_id", None)
                st.rerun()
        with ba4:
            if st.button(
                firma_lbl,
                disabled=firmar_disabled,
                use_container_width=True,
                key=f"fcd_fr_{aid_s}",
            ):
                _clear_calendar_dialog_focus()
                _open_firma_contrato_nav(dict(r), aid)

        if st.button("Cerrar", use_container_width=True, key=f"fcd_close_{aid_s}"):
            if form_dirty:
                st.session_state[f"fcd_confirm_leave_{aid_s}"] = True
                st.rerun()
            st.session_state.pop(seed_k, None)
            _clear_calendar_dialog_focus()
            st.rerun()
    else:
        if st.button("Cerrar", use_container_width=True, key=f"fcd_close_ro_{aid_s}"):
            st.session_state.pop(seed_k, None)
            _clear_calendar_dialog_focus()
            st.rerun()


@st.dialog("Citas del día", width="large", dismissible=False)
def _dialog_calendar_day_appointments(
    buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
    hist_counts: dict[str, int],
) -> None:
    tup = st.session_state.get("_cal_overflow_day")
    if not tup:
        return
    y, m, d = int(tup[0]), int(tup[1]), int(tup[2])
    day_rows = list(buckets.get((y, m, d), []))
    day_date = date(y, m, d)
    if not day_rows:
        st.info("No hay citas para este día con los filtros actuales.")
        if st.button("Cerrar", key="cal_dlg_close_empty", use_container_width=True):
            _clear_calendar_dialog_focus()
            st.rerun()
        return
    st.markdown(f"**{day_date.strftime('%d/%m/%Y')}** · **{len(day_rows)}** cita(s)")
    if _panel_is_technician_role():
        st.caption(
            "Como **tatuador / perforador** solo ves citas **desde hoy** con estado activo; aquí solo puedes "
            "**Completar firma profesional** cuando recepción ya guardó el contrato del cliente. "
            "El agendamiento, montos y reprogramación los gestiona administración o ventas."
        )
    else:
        st.caption(
            "Resumen de cada bloque **Cita**. Abre **Ficha completa** para ver datos del cliente, horario, montos y abonos."
        )
    for idx, r in enumerate(day_rows):
        appt_id = int(r.get("id", 0) or 0)
        _calendar_overflow_day_sheet_link_row(
            r,
            hist_counts,
            key_suffix=f"{appt_id}_{y}_{m}_{d}_{idx}",
        )
        if idx < len(day_rows) - 1:
            st.divider()
    if st.button("Cerrar", key="cal_dlg_close", use_container_width=True):
        _clear_calendar_dialog_focus()
        st.rerun()


@st.dialog("Cita", width="large", dismissible=False)
def _dialog_calendar_single_appointment(
    _buckets: dict[tuple[int, int, int], list[dict[str, Any]]],
    hist_counts: dict[str, int],
) -> None:
    """Diálogo para una sola cita (desde ▸ en rejilla semanal o mes)."""
    raw_id = st.session_state.get("_cal_focus_appt_id")
    try:
        aid = int(raw_id)
    except (TypeError, ValueError):
        aid = 0
    if aid <= 0:
        return
    r = _find_appointment_row_by_id(aid)
    if not r:
        st.warning(
            "No se encontró la cita en la lista actual. Pulsa **Actualizar** o abre de nuevo desde el calendario."
        )
        if st.button("Cerrar", key="cal_single_close_nf", use_container_width=True):
            _clear_calendar_dialog_focus()
            st.rerun()
        return
    try:
        day_date = _parse_date(r.get("appointment_date", r.get("date")))
    except (TypeError, ValueError):
        day_date = date.today()
    st.markdown(f"**{day_date.strftime('%d/%m/%Y')}** · Cita **#{aid}**")
    if _panel_is_technician_role():
        st.caption(
            "Como **tatuador / perforador** solo ves citas **desde hoy** con estado activo; aquí solo puedes "
            "**Completar firma profesional** cuando recepción ya guardó el contrato del cliente."
        )
    else:
        st.caption("Edita horario (reprogramación), prioridad (pill), artista y abonos. **Guardar** aplica contra la API.")
    _render_calendar_focus_appointment_body(r, hist_counts)


_AP_FIN_PAYMENTS_CACHE_PREFIX = "_ap_fin_payments_"
_AP_RECEIPTS_CACHE_PREFIX = "_ap_receipts_list_"


def _purge_appointment_payment_caches() -> None:
    """Evita historial de abonos obsoleto tras refrescar la lista de citas."""
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and k.startswith(_AP_FIN_PAYMENTS_CACHE_PREFIX):
            st.session_state.pop(k, None)


def _purge_appointment_receipt_caches() -> None:
    for k in list(st.session_state.keys()):
        if not isinstance(k, str):
            continue
        if k.startswith(_AP_RECEIPTS_CACHE_PREFIX) or k.startswith("_ap_receipt_pdf_"):
            st.session_state.pop(k, None)


def _get_appointment_payments_cached(appt_id: int) -> tuple[bool, int, Any]:
    """Un GET por cita y sesión (se invalida al refrescar citas o tras guardar montos)."""
    key = f"{_AP_FIN_PAYMENTS_CACHE_PREFIX}{int(appt_id)}"
    hit = st.session_state.get(key)
    if isinstance(hit, tuple) and len(hit) == 3:
        return hit[0], hit[1], hit[2]
    with st.spinner("Cargando historial de abonos…"):
        ok_p, code_p, payments = api_client.get_appointment_payments(appt_id)
    st.session_state[key] = (ok_p, code_p, payments)
    return ok_p, code_p, payments


def _fetch_appointments() -> None:
    qid = _appointments_query_assigned_user_id()
    ok, code, data = api_client.get_appointments(assigned_panel_user_id=qid)
    if ok and isinstance(data, list):
        data = _filter_appointments_for_session_role(data)
        st.session_state["_ap_list"] = data
        st.session_state["_ap_err"] = None
        _purge_appointment_payment_caches()
        _purge_appointment_receipt_caches()
    else:
        st.session_state["_ap_list"] = []
        st.session_state["_ap_err"] = f"HTTP {code}: {_api_error(data)}"





def _render_cita_row_actions(r: Dict[str, Any], *, show_firma: bool = True) -> None:
    """
    Menú de acciones por fila. `show_firma=False` omite la firma en vista administrativa (p. ej. Reporte).

    Tatuadores y perforadores solo ven acciones de **firma profesional** (sin flujo de cliente ni encuesta).
    """
    appt_id = int(r.get("id", 0) or 0)
    status = str(r.get("status") or "Agendada")
    has_customer = r.get("customer_id") is not None
    firmar_disabled = _firmar_contrato_disabled(r)
    repro_disabled = _reprogram_disabled_for_row(r)
    montos_disabled = appt_id <= 0 or status not in {"Agendada", "Reprogramada"}
    anular_disabled = appt_id <= 0 or status in {"Cancelada", "Finalizada"}
    firma_lbl = _firmar_contrato_button_label(r)

    if _panel_is_technician_role():
        pop = getattr(st, "popover", None)
        if pop:
            with pop("Firma profesional", use_container_width=True):
                if appt_id > 0:
                    st.caption(f"Cita #{appt_id}")
                if st.button(
                    firma_lbl,
                    disabled=firmar_disabled,
                    use_container_width=True,
                    key=f"pop_firmar_{appt_id}",
                ):
                    _open_firma_contrato_nav(r, appt_id)
            return
        if st.button(
            firma_lbl,
            disabled=firmar_disabled,
            use_container_width=True,
            key=f"fb_tech_only_{appt_id}",
        ):
            _open_firma_contrato_nav(r, appt_id)
        return

    pop = getattr(st, "popover", None)
    if pop:
        with pop("Acciones", use_container_width=True):
            if appt_id > 0:
                st.caption(f"Cita #{appt_id}")
                st.caption(f"Artista: **{_assigned_artist_display_name(r)}**")
            if show_firma:
                if st.button(
                    "Firmar contrato",
                    disabled=firmar_disabled,
                    use_container_width=True,
                    key=f"pop_firmar_{appt_id}",
                ):
                    _open_firma_contrato_nav(r, appt_id)
            if st.button(
                "Reprogramar cita",
                disabled=repro_disabled,
                use_container_width=True,
                key=f"pop_repr_{appt_id}",
                help="Solo **Agendada** o **Reprogramada** y sin contrato firmado. Tras firmar, la cita queda finalizada y no se reprograma.",
            ):
                st.session_state["_ap_reprogram_item"] = r
                st.rerun()
            if st.button(
                "Montos",
                disabled=montos_disabled,
                use_container_width=True,
                key=f"pop_fin_{appt_id}",
            ):
                st.session_state["_ap_fin_item"] = r
                st.rerun()
            rec_dis = appt_id <= 0
            if st.button(
                "Recibos (PDF)",
                disabled=rec_dis,
                use_container_width=True,
                key=f"pop_rec_{appt_id}",
                help="Descargar comprobantes emitidos al agendar o al abonar",
            ):
                st.session_state["_ap_receipts_item"] = r
                st.rerun()
            if st.button(
                "Anular",
                disabled=anular_disabled,
                use_container_width=True,
                key=f"pop_can_{appt_id}",
            ):
                st.session_state["_ap_cancel_item"] = r
                st.rerun()
        return

    if appt_id > 0:
        st.caption(f"Cita #{appt_id}")
    st.caption(f"Artista: **{_assigned_artist_display_name(r)}**")
    if show_firma:
        ln1, ln2 = st.columns(2)
        with ln1:
            if st.button(
                "Firmar",
                disabled=firmar_disabled,
                use_container_width=True,
                key=f"fb_compact_{appt_id}",
            ):
                _open_firma_contrato_nav(r, appt_id)
        with ln2:
            if st.button("Mover", disabled=repro_disabled, use_container_width=True, key=f"fb_repr_{appt_id}"):
                st.session_state["_ap_reprogram_item"] = r
                st.rerun()
    else:
        if st.button("Mover", disabled=repro_disabled, use_container_width=True, key=f"fb_repr_{appt_id}"):
            st.session_state["_ap_reprogram_item"] = r
            st.rerun()
    bn1, bn2, bn3 = st.columns(3)
    with bn1:
        if st.button("Montos", disabled=montos_disabled, use_container_width=True, key=f"fb_fin_{appt_id}"):
            st.session_state["_ap_fin_item"] = r
            st.rerun()
    with bn2:
        rec_key_dis = appt_id <= 0
        if st.button(
            "Recibos",
            disabled=rec_key_dis,
            use_container_width=True,
            key=f"fb_rec_{appt_id}",
            help="PDF si hubo abono al agendar y por cada abono adicional",
        ):
            st.session_state["_ap_receipts_item"] = r
            st.rerun()
    with bn3:
        if st.button("Anular", disabled=anular_disabled, use_container_width=True, key=f"fb_can_{appt_id}"):
            st.session_state["_ap_cancel_item"] = r
            st.rerun()


def _apply_appointment_filters(
    items: list[dict[str, Any]],
    *,
    use_date_range: bool = True,
    name_key: str = "_ap_f_name",
    service_key: str = "_ap_f_service",
    status_key: str = "_ap_f_status",
) -> list[dict[str, Any]]:
    text = str(st.session_state.get(name_key) or "").strip().lower()
    svc = str(st.session_state.get(service_key) or "Todos")
    status = str(st.session_state.get(status_key) or "Todos")
    from_date = st.session_state.get("_ap_f_from") if use_date_range else None
    to_date = st.session_state.get("_ap_f_to") if use_date_range else None
    filtered: list[dict[str, Any]] = []
    for row in items:
        name_value = str(row.get("customer_name", row.get("name", "")) or "")
        service_value = str(row.get("service_type", row.get("service", "")) or "")
        status_value = str(row.get("status") or "Agendada")
        appt_date = _parse_date(row.get("appointment_date", row.get("date")))
        if text and text not in name_value.lower():
            continue
        if svc != "Todos" and service_value != svc:
            continue
        if status != "Todos" and status_value != status:
            continue
        if from_date and appt_date < from_date:
            continue
        if to_date and appt_date > to_date:
            continue
        filtered.append(row)
    return filtered


def _cleanup_reprogram_dialog_state() -> None:
    keys = ("_ap_reprogram_seed_appt_id", "ap_reprogram_date", "ap_reprogram_slot", "ap_reprogram_detail")
    for k in keys:
        st.session_state.pop(k, None)


@st.dialog("Reprogramar cita", width="medium", dismissible=False)
def _dialog_reprogramar_cita() -> None:
    appt = st.session_state.get("_ap_reprogram_item") or {}
    appt_id = int(appt.get("id", 0) or 0)
    if appt_id <= 0:
        st.error("No se encontró la cita a reprogramar.")
        if st.button("Cerrar", use_container_width=True):
            st.session_state.pop("_ap_reprogram_item", None)
            _cleanup_reprogram_dialog_state()
            st.rerun()
        return
    if _reprogram_disabled_for_row(appt):
        st.warning(
            "No se puede reprogramar esta cita: debe estar **Agendada** o **Reprogramada**, "
            "sin **contrato firmado** y no cancelada."
        )
        if st.button("Cerrar", use_container_width=True, key="ap_reprogram_blocked_close"):
            st.session_state.pop("_ap_reprogram_item", None)
            _cleanup_reprogram_dialog_state()
            st.rerun()
        return
    seed_key = "_ap_reprogram_seed_appt_id"
    detail_default = str(appt.get("detail") or "")
    _, max_date_appt = _date_range_100y_window()
    # Una sola fuente de verdad: session_state por key — evita value+key (provoca glitch del popover Calendar)
    if st.session_state.get(seed_key) != appt_id:
        st.session_state[seed_key] = appt_id
        d0, sl0 = _parse_existing_slot(appt.get("appointment_date", appt.get("date")))
        today_d = date.today()
        st.session_state["ap_reprogram_date"] = d0 if d0 >= today_d else today_d
        st.session_state["ap_reprogram_slot"] = sl0
        st.session_state["ap_reprogram_detail"] = detail_default

    st.caption(
        f"Cita #{appt_id} · {appt.get('customer_name', appt.get('name', ''))} · "
        f"Artista: **{_assigned_artist_display_name(appt)}**"
    )
    # Detalle primero para no autofocos en el calendar al abrir el diálogo
    new_detail = st.text_area(
        "Detalle actualizado (opcional)",
        height=90,
        key="ap_reprogram_detail",
    )
    today_d = date.today()
    new_date = st.date_input(
        "Nueva fecha de cita",
        min_value=today_d,
        max_value=max_date_appt,
        key="ap_reprogram_date",
        format="DD/MM/YYYY",
    )
    slot_opts = _time_slot_options()
    need_slots_repr = _duration_slots_for_existing_appointment(appt)
    raw_list_repr = list(st.session_state.get("_ap_list") or [])
    sched_repr = appointment_to_contract_kind(appt)
    ra_raw = appt.get("assigned_panel_user_id")
    artist_repr: Optional[int] = None
    if ra_raw not in (None, "", 0):
        try:
            artist_repr = int(ra_raw)
        except (TypeError, ValueError):
            artist_repr = None
    if artist_repr is not None:
        day_rows_repr = _appointments_for_artist_schedule(
            raw_list_repr,
            new_date,
            artist_repr,
            schedule_kind=sched_repr,
            exclude_appointment_id=appt_id,
        )
        st.caption(
            "Franjas según **este profesional** y solo citas del **mismo tipo** (tatuaje o piercing)."
        )
    else:
        day_rows_repr = _appointments_same_day_schedule_kind(
            raw_list_repr, new_date, sched_repr
        )
        st.caption(
            "Sin profesional asignado en base de datos; se usan citas del **mismo tipo** ese día."
        )
    busy_repr = _busy_slot_indices_for_day(day_rows_repr, slot_opts)
    avail_repr = _available_start_slots(slot_opts, need_slots_repr, busy_repr)
    if not avail_repr:
        st.warning(
            "No hay franjas libres ese día para esta duración. Puedes forzar una hora de la lista completa abajo; revisa conflictos en agenda."
        )
        avail_repr = slot_opts
    cur_sl = st.session_state.get("ap_reprogram_slot")
    if cur_sl not in avail_repr:
        st.session_state["ap_reprogram_slot"] = avail_repr[0]
    new_slot = st.selectbox(
        "Nueva franja horaria *",
        options=avail_repr,
        key="ap_reprogram_slot",
    )
    dt_reschedule = _combine_appointment_datetime(new_date, str(new_slot))
    c1, c2 = st.columns(2)
    with c1:
        if st.button(
            "Guardar reprogramación",
            type="primary",
            use_container_width=True,
            key="ap_reprogram_save_btn",
        ):
            with st.spinner("Aplicando reprogramación…"):
                ok, code, data = api_client.patch_appointment_reschedule(
                    appt_id,
                    dt_reschedule,
                    (new_detail or "").strip() or None,
                )
            if ok:
                pretty = _format_dt_for_user_message(dt_reschedule)
                _queue_appointment_action_success(
                    f"**Cita reprogramada** · #{appt_id} · nueva fecha y hora: **{pretty}**."
                )
                st.session_state["_ap_reload"] = True
                st.session_state.pop("_ap_reprogram_item", None)
                _cleanup_reprogram_dialog_state()
                st.rerun()
            else:
                st.toast(
                    f"Error HTTP {code}: {_api_error(data)}",
                    icon="❌",
                    duration="long",
                )
    with c2:
        if st.button("Cancelar", use_container_width=True, key="ap_reprogram_close_btn"):
            st.session_state.pop("_ap_reprogram_item", None)
            _cleanup_reprogram_dialog_state()
            st.rerun()


def _label_cancel_abono(v: str) -> str:
    if v == "credito_cliente":
        return "Saldo a favor del cliente — el abono pasa a crédito interno y deja de contar como cobrado sobre la cita"
    return "Devolución — el abono deja la cita como no cobrado (sin saldo a favor aquí)"

@st.dialog("Confirmar anulación", width="medium", dismissible=False)
def _dialog_cancelar_cita() -> None:
    appt = st.session_state.get("_ap_cancel_item") or {}
    appt_id = int(appt.get("id", 0) or 0)
    if appt_id <= 0:
        st.error("No se encontró la cita a anular.")
        if st.button("Cerrar", use_container_width=True, key="ap_cancel_close_missing"):
            st.session_state.pop("_ap_cancel_item", None)
            st.rerun()
        return
    deposit = float(appt.get("deposit") or 0)
    art_nm = _assigned_artist_display_name(appt)
    warning = (
        f"Vas a anular la cita #{appt_id} de "
        f"{appt.get('customer_name', appt.get('name', 'cliente'))}. "
        f"Artista asignado: **{art_nm}**. Esta acción cambia el estado a Cancelada."
    )
    if deposit > 0:
        warning += f" Hay {_format_cop(deposit)} abonados en esta fila."
    else:
        warning += " No hay abonos registrados en esta cita."
    st.warning(warning)

    cancel_abono: str
    if deposit > 0:
        st.markdown("Si hubo abono, cómo debe reflejarse para **resumen y totales**:", unsafe_allow_html=True)
        cancel_abono = st.radio(
            "Tratamiento del abono",
            ("credito_cliente", "devolucion"),
            format_func=_label_cancel_abono,
            horizontal=False,
            key=f"dlg_cancel_abono_radio_{appt_id}",
            label_visibility="visible",
        )
    else:
        cancel_abono = "devolucion"
        st.caption("Sin abono; la anulación solo cierra la cita en el sistema.")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Sí, anular", type="primary", use_container_width=True, key="ap_cancel_confirm_btn"):
            with st.spinner("Anulando cita…"):
                ok, code, data = api_client.patch_appointment_status(appt_id, "Cancelada", cancel_abono)
            if ok:
                _queue_appointment_action_success(
                    f"**Cita anulada** · #{appt_id} · estado **Cancelada**."
                )
                st.session_state["_ap_reload"] = True
                st.session_state.pop("_ap_cancel_item", None)
                st.rerun()
            else:
                st.toast(
                    f"Error HTTP {code}: {_api_error(data)}",
                    icon="❌",
                    duration="long",
                )
    with c2:
        if st.button("No, volver", use_container_width=True, key="ap_cancel_back_btn"):
            st.session_state.pop("_ap_cancel_item", None)
            st.rerun()


@st.dialog("Ajustar montos", width="medium", dismissible=False)
def _dialog_ajustar_montos() -> None:
    appt = st.session_state.get("_ap_fin_item") or {}
    appt_id = int(appt.get("id", 0) or 0)
    status = str(appt.get("status") or "Agendada")
    if appt_id <= 0:
        st.error("No se encontró la cita.")
        if st.button("Cerrar", use_container_width=True, key="ap_fin_close_missing"):
            st.session_state.pop("_ap_fin_item", None)
            st.rerun()
        return
    if status not in {"Agendada", "Reprogramada"}:
        st.error("Solo puedes editar montos en estados Agendada o Reprogramada.")
        if st.button("Cerrar", use_container_width=True, key="ap_fin_close_invalid"):
            st.session_state.pop("_ap_fin_item", None)
            st.rerun()
        return
    st.caption(
        f"Cita #{appt_id} · Estado: {status} · Artista: **{_assigned_artist_display_name(appt)}**"
    )

    if st.session_state.get("_ap_fin_dialog_appt_id") != appt_id:
        st.session_state.pop("_ap_fin_save_error", None)
    st.session_state["_ap_fin_dialog_appt_id"] = appt_id

    st.markdown("##### Historial de abonos")
    ok_p, code_p, payments = _get_appointment_payments_cached(appt_id)
    if ok_p and isinstance(payments, list):
        if payments:
            for p in payments:
                when = str(p.get("created_at") or "")
                note = str(p.get("note") or "Sin nota")
                amount = _to_float(p.get("amount"), 0.0)
                st.write(f"- {when[:19]} · {_format_cop(amount)} · {note}")
        else:
            st.info("Aún no hay abonos registrados.")
    else:
        st.warning(f"No se pudo cargar historial (HTTP {code_p}).")

    current_total = float(appt.get("total_amount") or 0)
    current_deposit = float(appt.get("deposit") or 0)
    total_amount = st.number_input(
        "Valor total del trabajo (COP)",
        min_value=0.0,
        step=10000.0,
        value=current_total,
        key="ap_fin_total",
    )
    pending = round(float(total_amount) - float(current_deposit), 2)
    st.caption(f"Abonado actual: {_format_cop(current_deposit)}")
    st.caption(f"Saldo pendiente calculado: {_format_cop(max(pending, 0))}")

    pend_ui = max(float(pending), 0.0)
    can_add_extra = pend_ui > 0.009
    if not can_add_extra:
        st.info("Trabajo cubierto: no hay saldo pendiente; no se pueden agregar abonos adicionales.")
        st.session_state["ap_fin_extra_payment"] = 0.0
        st.session_state["ap_fin_extra_note"] = ""

    extra_payment = st.number_input(
        "Agregar abono adicional (COP)",
        min_value=0.0,
        max_value=float(pend_ui) if can_add_extra else 0.0,
        step=10000.0,
        key="ap_fin_extra_payment",
        disabled=not can_add_extra,
        help=(
            "Solo si el saldo pendiente es mayor a cero."
            if can_add_extra
            else "Saldo pendiente en cero; no aplica otro abono."
        ),
    )
    payment_note = st.text_input(
        "Nota del abono (opcional)",
        key="ap_fin_extra_note",
        placeholder="Ej: abono en efectivo",
        disabled=not can_add_extra,
    )

    save_err = st.session_state.get("_ap_fin_save_error")
    _toast_financial_save_error_if_any()

    c1, c2 = st.columns(2)
    with c1:
        do_save = st.button("Guardar", type="primary", use_container_width=True, key="ap_fin_save_btn")
    with c2:
        do_cancel = st.button("Cancelar", use_container_width=True, key="ap_fin_cancel_btn")

    if save_err:
        if st.button("Cerrar", use_container_width=True, key="ap_fin_err_close"):
            st.session_state.pop("_ap_fin_save_error", None)
            with st.spinner("Cerrando…"):
                st.session_state.pop("_ap_fin_item", None)
                st.session_state.pop("ap_fin_total", None)
                st.session_state.pop("ap_fin_extra_payment", None)
                st.session_state.pop("ap_fin_extra_note", None)
                st.session_state.pop("_ap_fin_dialog_appt_id", None)
            st.rerun()

    if do_cancel:
        st.session_state.pop("_ap_fin_save_error", None)
        with st.spinner("Cerrando…"):
            st.session_state.pop("_ap_fin_item", None)
            st.session_state.pop("ap_fin_total", None)
            st.session_state.pop("ap_fin_extra_payment", None)
            st.session_state.pop("ap_fin_extra_note", None)
            st.session_state.pop("_ap_fin_dialog_appt_id", None)
        st.rerun()

    if do_save:
        if current_deposit > total_amount:
            st.session_state["_ap_fin_save_error"] = (
                "El abonado acumulado no puede ser mayor al valor total."
            )
            st.rerun()
        ex = float(st.session_state.get("ap_fin_extra_payment") or 0)
        if ex > 0 and not can_add_extra:
            st.session_state["_ap_fin_save_error"] = (
                "No hay saldo pendiente; no puedes registrar otro abono."
            )
            st.rerun()
        err_save: Optional[str] = None
        with st.spinner("Guardando montos y abonos…"):
            ok, code, data = api_client.patch_appointment_financials(
                appt_id,
                float(total_amount),
                float(current_deposit),
                float(max(pending, 0)),
            )
            if not ok:
                err_save = f"Error HTTP {code}: {_api_error(data)}"
            elif ex > 0:
                note_s = (st.session_state.get("ap_fin_extra_note") or "").strip()
                ok_pay, code_pay, data_pay = api_client.post_appointment_payment(
                    appt_id,
                    ex,
                    note_s or None,
                )
                if not ok_pay:
                    err_save = f"No se pudo registrar abono (HTTP {code_pay}): {_api_error(data_pay)}"
        if err_save:
            st.session_state["_ap_fin_save_error"] = err_save
            st.rerun()
        st.session_state.pop("_ap_fin_save_error", None)
        st.session_state.pop(f"{_AP_FIN_PAYMENTS_CACHE_PREFIX}{appt_id}", None)
        _purge_appointment_receipt_caches()
        if ex > 0:
            _queue_appointment_action_success(
                "**Montos y abonos actualizados.** Hay un nuevo recibo PDF en **Recibos**."
            )
        else:
            _queue_appointment_action_success(
                "**Montos actualizados** (valor total del trabajo y saldos)."
            )
        st.session_state["_ap_reload"] = True
        st.session_state.pop("_ap_fin_item", None)
        st.session_state.pop("ap_fin_total", None)
        st.session_state.pop("ap_fin_extra_payment", None)
        st.session_state.pop("ap_fin_extra_note", None)
        st.session_state.pop("_ap_fin_dialog_appt_id", None)
        st.rerun()


@st.dialog("Recibos de pago (PDF)", width="large", dismissible=False)
def _dialog_recibos_cita() -> None:
    appt = st.session_state.get("_ap_receipts_item") or {}
    appt_id = int(appt.get("id", 0) or 0)
    if appt_id <= 0:
        st.error("No se encontró la cita.")
        if st.button("Cerrar", use_container_width=True, key="ap_rec_close_bad"):
            st.session_state.pop("_ap_receipts_item", None)
            st.rerun()
        return
    name = str(appt.get("customer_name") or appt.get("name") or "").strip()
    st.markdown(f"**Cita #{appt_id}** · {name or '—'}")
    st.caption(
        "Si al crear la cita hubo abono y el servicio es **tatuaje** (u otro no piercing), se genera un recibo inicial; "
        "en **piercing / limpieza / cambio** no se envía recibo PDF al agendar (solo notificación de cita). "
        "Cada abono adicional puede generar otro PDF. Los archivos se guardan en el servidor."
    )

    list_key = f"{_AP_RECEIPTS_CACHE_PREFIX}{appt_id}"
    cached = st.session_state.get(list_key)
    if not isinstance(cached, tuple) or len(cached) != 3:
        with st.spinner("Cargando índice de recibos…"):
            ok, code, data = api_client.get_appointment_receipts(appt_id)
        st.session_state[list_key] = (ok, code, data)
    ok, code, data = st.session_state[list_key]
    if not ok:
        st.error(f"No se pudieron listar los recibos (HTTP {code}): {_api_error(data)}")
        if st.button("Cerrar", use_container_width=True, key="ap_rec_close_list_err"):
            st.session_state.pop(list_key, None)
            st.session_state.pop("_ap_receipts_item", None)
            st.rerun()
        return

    rows: List[Dict[str, Any]] = []
    if isinstance(data, list):
        rows = [x for x in data if isinstance(x, dict)]

    if not rows:
        st.info("Todavía no hay recibos registrados para esta cita.")

    for r in rows:
        rid = int(r.get("id", 0) or 0)
        if rid <= 0:
            continue
        kind = str(r.get("kind") or "")
        kind_es = "Agenda / primer abono" if kind == "inicial" else "Abono adicional"
        try:
            amt = float(r.get("amount") or 0)
        except (TypeError, ValueError):
            amt = 0.0
        when = str(r.get("created_at") or "")
        if len(when) >= 19:
            when = when[:19]
        st.markdown(f"**{kind_es}** · {when or '—'} · **{_format_cop(amt)}**")

        pdf_key = f"_ap_receipt_pdf_{appt_id}_{rid}"
        if pdf_key not in st.session_state:
            ok_pdf, _pc, blob, fname = api_client.fetch_appointment_receipt_pdf(appt_id, rid)
            if ok_pdf and blob:
                st.session_state[pdf_key] = (blob, fname)
        got = st.session_state.get(pdf_key)
        if isinstance(got, tuple) and len(got) == 2 and got[0]:
            blob, fname = got[0], got[1]
            st.download_button(
                "Descargar PDF",
                data=blob,
                file_name=str(fname or f"recibo_{appt_id}_{rid}.pdf"),
                mime="application/pdf",
                use_container_width=True,
                key=f"ap_rec_dl_{appt_id}_{rid}",
            )
        else:
            st.caption("No se pudo cargar el archivo PDF.")
        st.divider()

    if st.button("Cerrar", use_container_width=True, key="ap_rec_close_main"):
        st.session_state.pop("_ap_receipts_item", None)
        st.rerun()




def _inject_citas_shared_styles() -> None:
    """CSS consolidado desde `streamlit_app/styles/` (arquitectura del tab Citas)."""
    inject_via_streamlit_lazy()


def _init_appt_tab_session_state() -> None:
    if "_ap_page" not in st.session_state:
        st.session_state["_ap_page"] = 0
    if "_ap_limit" not in st.session_state:
        st.session_state["_ap_limit"] = 10
    if "_ap_reload" not in st.session_state:
        st.session_state["_ap_reload"] = True
    if "_ap_f_name" not in st.session_state:
        st.session_state["_ap_f_name"] = ""
    if "_ap_f_service" not in st.session_state:
        st.session_state["_ap_f_service"] = "Todos"
    if "_ap_f_status" not in st.session_state:
        st.session_state["_ap_f_status"] = "Todos"
    if "_ap_f_from" not in st.session_state:
        st.session_state["_ap_f_from"] = None
    if "_ap_f_to" not in st.session_state:
        st.session_state["_ap_f_to"] = None
    if "_ap_cal_f_name" not in st.session_state:
        st.session_state["_ap_cal_f_name"] = ""
    if "_ap_cal_f_service" not in st.session_state:
        st.session_state["_ap_cal_f_service"] = "Todos"
    if "_ap_cal_f_status" not in st.session_state:
        st.session_state["_ap_cal_f_status"] = "Todos"


def _render_procedure_value_bar_chart(filtered_items: list[dict[str, Any]]) -> None:
    """Barras: suma de total trabajo por tipo de servicio (procedimiento), según el filtro actual."""
    if not filtered_items:
        return
    by_svc: dict[str, float] = defaultdict(float)
    for row in filtered_items:
        svc = str(row.get("service_type", row.get("service", "")) or "").strip() or "Sin especificar"
        t_total, _, _ = _financial_row_values(row)
        by_svc[svc] += t_total
    ordered = sorted(by_svc.items(), key=lambda x: -x[1])
    categories = [k for k, _ in ordered]
    values = [float(v) for _, v in ordered]
    st.markdown("##### Valor por procedimiento")
    report_charts.render_vertical_bars(
        st,
        categories=categories,
        values=values,
        x_title="Tipo de servicio / procedimiento",
        y_title="Total trabajo (COP)",
        height=min(420, 140 + len(ordered) * 42),
        hovertemplate="<b>%{x}</b><br>%{y:,.0f} COP<extra></extra>",
        key="rep_fin_valor_procedimiento",
    )


def _truncate_survey_chart_label(s: str, max_len: int = 50) -> str:
    t = str(s).replace("\n", " ").strip()
    if len(t) <= max_len:
        return t
    return t[: max_len - 1] + "…"


def _survey_pie_chart_from_counts(
    counts: dict[str, int],
    *,
    chart_key: str,
    sort_key: Optional[Any] = None,
    reverse: bool = False,
    limit: Optional[int] = None,
) -> None:
    """Gráfica de torta (Plotly; mismo tema que barras del reporte)."""
    if not counts:
        return
    items = [(str(k), int(v)) for k, v in counts.items() if int(v) > 0]
    if not items:
        return
    if sort_key is not None:
        items.sort(key=sort_key, reverse=reverse)
    else:
        items.sort(key=lambda x: x[1], reverse=True)
    if limit is not None and limit > 0 and len(items) > limit:
        head = list(items[: max(0, limit - 1)])
        tail = items[limit - 1 :]
        otros = sum(v for _, v in tail)
        if otros > 0:
            head.append(("Otros", otros))
        items = head
    pie_labels = [_truncate_survey_chart_label(k) for k, _ in items]
    pie_values = [v for _, v in items]
    if sum(pie_values) <= 0:
        return
    report_charts.render_pie(st, labels=pie_labels, values=pie_values, height=440, key=chart_key)


def _normalize_survey_label_ascii_lower(s: str) -> str:
    """Compara etiquetas sin distinguir tildes ni mayúsculas."""
    t = unicodedata.normalize("NFKD", str(s or ""))
    return "".join(c for c in t if not unicodedata.combining(c)).lower()


def _survey_question_is_procedure_value_question(label: str) -> bool:
    """P. ej. «¿Cuánto es el valor de tu procedimiento?» → barras Plotly (mismo estilo que finanzas)."""
    n = _normalize_survey_label_ascii_lower(label)
    return "procedimiento" in n and "valor" in n


def _pairs_from_number_breakdown(nb: dict[str, int]) -> list[tuple[float, int]]:
    out: list[tuple[float, int]] = []
    for k, v in nb.items():
        try:
            out.append((float(k), int(v)))
        except (TypeError, ValueError):
            continue
    out.sort(key=lambda x: x[0])
    return out


def _survey_number_bar_chart_2d(pairs: list[tuple[float, int]], *, x_title: str, chart_key: str) -> None:
    """Barras respuesta numérica × frecuencia (Plotly, mismo estilo que el resto del reporte)."""
    if not pairs:
        return
    vals = [p[0] for p in pairs]
    ns = [p[1] for p in pairs]
    categories = [f"{v:g}" for v in vals]
    report_charts.render_vertical_bars(
        st,
        categories=categories,
        values=ns,
        x_title=x_title,
        y_title="Respuestas (n)",
        height=min(400, 140 + len(categories) * 36),
        hovertemplate="<b>Valor %{x}</b><br>%{y} respuesta(s)<extra></extra>",
        key=chart_key,
    )


def _render_survey_question_stats_report() -> None:
    ok, code, raw = get_survey_question_stats_summary_cached()
    if not ok:
        det = _api_error(raw)
        st.warning(
            f"No se pudieron cargar las estadísticas de encuesta (HTTP {code}). "
            f"Ejecuta las migraciones `011`–`014` en `sql/` según corresponda. Detalle: {det}"
        )
        return
    if not isinstance(raw, list) or len(raw) == 0:
        st.caption("No hay preguntas registradas o la lista está vacía.")
        return
    for idx, row in enumerate(raw):
        if not isinstance(row, dict):
            continue
        qid = int(row.get("question_id") or idx)
        label = str(row.get("label") or "")
        qt = str(row.get("question_type") or "")
        ql = question_type_label_es(qt)
        ck = SCOPE_LABEL_ES.get(str(row.get("contract_kind") or "tattoo"), "—")
        rc = int(row.get("response_count") or 0)
        supports_chart = question_type_supports_distribution_chart(qt)
        st.divider()
        st.markdown(f"**{label}** · _{ql}_ · **{ck}** · n = **{rc}**")
        chart_shown = False

        rb = row.get("rating_breakdown")
        if qt == "rating_1_5" and isinstance(rb, dict) and rb:
            def _rk(item: tuple[str, int]) -> int:
                try:
                    return int(item[0])
                except (TypeError, ValueError):
                    return 0

            _survey_pie_chart_from_counts(dict(rb), sort_key=_rk, chart_key=f"rep_survey_pie_{qid}_rating")
            chart_shown = True
            if row.get("avg_rating") is not None:
                st.metric("Promedio (1–5)", f"{float(row['avg_rating']):.2f}")
        elif qt == "yes_no":
            yc = int(row.get("yes_count") or 0)
            nc = int(row.get("no_count") or 0)
            c1, c2 = st.columns(2)
            c1.metric("Sí", yc)
            c2.metric("No", nc)
            if yc + nc > 0:
                _survey_pie_chart_from_counts(
                    {"Sí": yc, "No": nc},
                    sort_key=lambda x: 0 if x[0] == "Sí" else 1,
                    chart_key=f"rep_survey_pie_{qid}_yesno",
                )
                chart_shown = True
        elif qt == "number":
            nb = row.get("number_breakdown")
            if isinstance(nb, dict) and nb:

                def _nk(item: tuple[str, int]) -> float:
                    try:
                        return float(item[0])
                    except (TypeError, ValueError):
                        return 0.0

                pairs = _pairs_from_number_breakdown(dict(nb))
                if _survey_question_is_procedure_value_question(label) and pairs:
                    _survey_number_bar_chart_2d(
                        pairs,
                        x_title="Valor informado (tu procedimiento)",
                        chart_key=f"rep_survey_bar_{qid}_procval",
                    )
                    chart_shown = True
                else:
                    _survey_pie_chart_from_counts(dict(nb), sort_key=_nk, chart_key=f"rep_survey_pie_{qid}_number")
                    chart_shown = True
            if row.get("avg_number") is not None:
                st.metric("Promedio numérico", f"{float(row['avg_number']):.4f}")
        elif qt in ("radio", "select", "checkbox"):
            cb = row.get("choice_breakdown")
            if isinstance(cb, dict) and cb:
                lim = 24 if qt == "checkbox" else 32
                _survey_pie_chart_from_counts(dict(cb), limit=lim, chart_key=f"rep_survey_pie_{qid}_choice")
                chart_shown = True
                if qt == "checkbox":
                    st.caption(
                        "Casillas: cada **sector** puede ser una combinación guardada (texto/JSON); "
                        "no son opciones independientes. Si hay muchas categorías, el resto se agrupa en **Otros**."
                    )
        elif qt in ("text", "textarea", "text_short"):
            tc = int(row.get("text_response_count") or 0)
            st.caption(
                f"Pregunta de **texto libre**: no tiene categorías fijas adecuadas para una torta. "
                f"Respuestas no vacías: **{tc}**."
            )
        else:
            tc = int(row.get("text_response_count") or 0)
            st.caption(f"Respuestas con texto registrado: {tc}")

        if supports_chart and not chart_shown and rc > 0:
            st.info("Hay respuestas, pero aún no hay datos agregados para graficar (revisa el tipo de pregunta).")
        elif supports_chart and rc == 0:
            st.caption("Sin respuestas todavía.")


def _excel_fingerprint(rows: list[dict[str, Any]]) -> str:
    """Hash estable de los IDs de las filas filtradas (no del contenido completo)."""
    ids = sorted(int(r.get("id") or 0) for r in rows)
    return hashlib.md5(json.dumps(ids, separators=(",", ":")).encode()).hexdigest()[:16]


def _get_excel_cached(rows: list[dict[str, Any]]) -> bytes:
    """
    Excel del filtro actual desde session_state si el conjunto de IDs no cambió;
    lo regenera y cachea si no hay hit. Elimina otros buffers _ap_xlsx_* previos.
    """
    fp = _excel_fingerprint(rows)
    cache_key = f"_ap_xlsx_{fp}"
    hit = st.session_state.get(cache_key)
    if isinstance(hit, bytes) and len(hit) > 0:
        return hit

    for k in [
        k
        for k in st.session_state
        if isinstance(k, str) and k.startswith("_ap_xlsx_") and k != cache_key
    ]:
        st.session_state.pop(k, None)

    data = _citas_filtered_to_excel_bytes(rows, generated_at=datetime.now())
    st.session_state[cache_key] = data
    return data


def _render_reporte_financiero_citas_body(
    items: list[dict[str, Any]],
    svc_values: list[str],
    status_values: list[str],
) -> None:
    """Filtros, métricas, export Excel y tabla paginada (solo finanzas)."""
    st.markdown("##### Filtros")
    f1, f2, f3, f4, f5 = st.columns([1.3, 1.0, 1.0, 0.9, 0.9])
    with f1:
        st.text_input("Filtrar nombre", key="_ap_f_name", placeholder="Nombre cliente")
    with f2:
        st.selectbox("Servicio", options=["Todos", *svc_values], key="_ap_f_service")
    with f3:
        st.selectbox("Estado", options=["Todos", *status_values], key="_ap_f_status")
    with f4:
        st.date_input("Desde", key="_ap_f_from")
    with f5:
        st.date_input("Hasta", key="_ap_f_to")

    hist_counts_raw = st.session_state.get("_ap_hist_counts")
    hist_counts = dict(hist_counts_raw) if isinstance(hist_counts_raw, dict) else {}
    if not hist_counts and items:
        hist_counts = _appointment_counts_by_client(items)

    filtered_items = _apply_appointment_filters(items)

    total_trabajo = 0.0
    total_abonado = 0.0
    total_pendiente = 0.0
    total_credito_favor = 0.0
    for row in filtered_items:
        row_total, row_abonado, row_pendiente = _financial_row_values(row)
        total_trabajo += row_total
        total_abonado += row_abonado
        total_pendiente += row_pendiente
        total_credito_favor += _customer_credit_value(row)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total trabajo", _format_cop(total_trabajo))
    m2.metric("Total abonado", _format_cop(total_abonado))
    m3.metric("Total saldo pendiente", _format_cop(total_pendiente))
    m4.metric("Saldo a favor (filtro)", _format_cop(total_credito_favor))

    _render_procedure_value_bar_chart(filtered_items)

    _informe_dt = datetime.now()
    try:
        _xlsx_agenda = _get_excel_cached(filtered_items)
    except Exception as e:
        _xlsx_agenda = b""
        if filtered_items:
            st.warning(f"No se pudo generar el Excel. Instala `openpyxl` en el venv: {e}")
    _dl_left, _dl_right = st.columns([4, 1])
    with _dl_right:
        st.download_button(
            label="Descargar Excel",
            help="Exporta financiero del filtro actual (nombre cliente y montos; hoja resumen).",
            data=_xlsx_agenda,
            file_name=(
                "Informe-finanzas-citas-"
                f"{_informe_dt.strftime('%Y-%m-%d-%H%M')}.xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            disabled=len(filtered_items) == 0 or len(_xlsx_agenda) == 0,
            key="btn_reporte_fin_xlsx",
        )

    st.markdown("##### Listado de citas")
    total = len(filtered_items)
    limit = int(st.session_state["_ap_limit"])
    page = int(st.session_state["_ap_page"])
    total_pages = max(1, (total + limit - 1) // limit)
    if page >= total_pages:
        page = max(0, total_pages - 1)
        st.session_state["_ap_page"] = page
    start = page * limit
    rows = filtered_items[start : start + limit]

    colw = [1.48, 1.0, 0.92, 0.82, 0.78, 0.78, 0.92, 0.85, 0.76, 1.52]
    h1, h2, h3, h4, h5, h6, h7, h8, h9, h10 = st.columns(colw)
    h1.markdown('<span class="ap-col-title">Nombre</span>', unsafe_allow_html=True)
    h2.markdown('<span class="ap-col-title">Artista</span>', unsafe_allow_html=True)
    h3.markdown('<span class="ap-col-title">Servicio</span>', unsafe_allow_html=True)
    h4.markdown('<span class="ap-col-title">Fecha y hora</span>', unsafe_allow_html=True)
    h5.markdown('<span class="ap-col-title">Total</span>', unsafe_allow_html=True)
    h6.markdown('<span class="ap-col-title">Abonado</span>', unsafe_allow_html=True)
    h7.markdown('<span class="ap-col-title">Pendiente</span>', unsafe_allow_html=True)
    h8.markdown('<span class="ap-col-title">A favor</span>', unsafe_allow_html=True)
    h9.markdown('<span class="ap-col-title">Estado</span>', unsafe_allow_html=True)
    h10.markdown('<span class="ap-col-title">Acciones</span>', unsafe_allow_html=True)
    for r in rows:
        c1, c2, c3, c4, c5, c6, c7, c8, c9, c10 = st.columns(colw)
        c1.markdown(_customer_name_pill_html(r, hist_counts), unsafe_allow_html=True)
        c2.write(_assigned_artist_display_name(r))
        c3.write(r.get("service_type", r.get("service", "")))
        c4.write(_format_appt_when(r.get("appointment_date", r.get("date", ""))))
        total_amount, deposit_amount, pending_balance = _financial_row_values(r)
        credito = _customer_credit_value(r)
        c5.write(_format_cop(total_amount))
        c6.write(_format_cop(deposit_amount))
        c7.write(_format_cop(pending_balance))
        c8.write("—" if credito <= 0 else _format_cop(credito))
        status = str(r.get("status") or "Agendada")
        c9.markdown(_status_pill_html(status), unsafe_allow_html=True)
        with c10:
            _render_cita_row_actions(r, show_firma=False)

    p1, p2, p3 = st.columns([1, 1, 2.5])
    with p1:
        st.write("")
        if st.button("◀", disabled=page <= 0, use_container_width=True, key="rep_ap_page_prev"):
            st.session_state["_ap_page"] = max(0, page - 1)
            st.rerun()
    with p2:
        st.write("")
        if st.button("▶", disabled=(page + 1) * limit >= total if total else True, use_container_width=True, key="rep_ap_page_next"):
            st.session_state["_ap_page"] = page + 1
            st.rerun()
    with p3:
        st.write("")
        st.caption(f"Página {page + 1}/{total_pages} · Total filtrado: {total} cita(s)")


def warm_session_after_login(allowed_module_keys: frozenset[str]) -> None:
    """Precarga agenda en sesión tras iniciar sesión (usar dentro del spinner en main)."""
    if "citas" not in allowed_module_keys and "reporte" not in allowed_module_keys:
        return
    _init_appt_tab_session_state()
    st.session_state["_ap_reload"] = True
    _sync_appointments_from_api()


def _invoke_citas_tab_dialogs(
    by_day: dict[tuple[int, int, int], list[dict[str, Any]]],
    hist_counts: dict[str, int],
) -> None:
    """Invoca diálogos al inicio del flujo para que el overlay no quede al final del DOM."""
    _technician_clear_disallowed_dialog_states()
    if (
        st.session_state.get("_ap_fin_item")
        or st.session_state.get("_ap_reprogram_item")
        or st.session_state.get("_ap_receipts_item")
    ):
        _clear_calendar_dialog_focus()
    if st.session_state.get("_ap_reprogram_item"):
        _dialog_reprogramar_cita()
    if st.session_state.get("_ap_fin_item"):
        _dialog_ajustar_montos()
    if st.session_state.get("_ap_cancel_item"):
        _dialog_cancelar_cita()
    if st.session_state.get("_ap_receipts_item"):
        _dialog_recibos_cita()
    if st.session_state.get("_cal_focus_appt_id"):
        _dialog_calendar_single_appointment(by_day, hist_counts)
    elif st.session_state.get("_cal_overflow_day"):
        _dialog_calendar_day_appointments(by_day, hist_counts)
    if st.session_state.get("_ap_dlg") == "create":
        _dialog_agendar_cita()


def _sync_appointments_from_api() -> None:
    """
    GET /appointments solo si hubo cambios, cambio de filtro o primera carga.
    Precalcula hist_counts y svc_values cuando se refrescan datos.
    """
    qid = _appointments_query_assigned_user_id()
    prev_qid = st.session_state.get("_ap_last_fetch_qid")
    fetch_needed = (
        st.session_state.get("_ap_reload", True)
        or prev_qid != qid
        or st.session_state.pop("_ap_refresh_after_contract", False)
    )
    if not fetch_needed:
        items_miss = st.session_state.get("_ap_list") or []
        if items_miss:
            if "_ap_hist_counts" not in st.session_state:
                st.session_state["_ap_hist_counts"] = _appointment_counts_by_client(items_miss)
            if "_ap_svc_values" not in st.session_state:
                st.session_state["_ap_svc_values"] = sorted(
                    {
                        str(i.get("service_type", i.get("service", "")) or "").strip()
                        for i in items_miss
                        if str(i.get("service_type", i.get("service", "")) or "").strip()
                    }
                )
        return

    with st.spinner("Actualizando citas…"):
        _fetch_appointments()

    items_post: list[Any] = st.session_state.get("_ap_list") or []
    st.session_state["_ap_hist_counts"] = _appointment_counts_by_client(items_post)
    st.session_state["_ap_svc_values"] = sorted(
        {
            str(i.get("service_type", i.get("service", "")) or "").strip()
            for i in items_post
            if str(i.get("service_type", i.get("service", "")) or "").strip()
        }
    )

    for k in [k for k in st.session_state if isinstance(k, str) and k.startswith("_ap_xlsx_")]:
        st.session_state.pop(k, None)

    st.session_state["_ap_reload"] = False
    st.session_state["_ap_last_fetch_qid"] = qid


def render_reporte_citas_tab() -> None:
    """Pestaña Reporte: finanzas y encuestas en sub-secciones; mismos filtros de citas para finanzas."""
    _init_appt_tab_session_state()
    _inject_citas_shared_styles()
    _sync_appointments_from_api()

    _technician_clear_disallowed_dialog_states()

    _render_appointments_fetch_error_toast()
    _render_appointment_action_feedback()

    items = list(st.session_state.get("_ap_list") or [])
    svc_values = list(st.session_state.get("_ap_svc_values") or [])
    status_values = ["Agendada", "Reprogramada", "Finalizada", "Cancelada"]

    if st.session_state.get("_ap_reprogram_item"):
        _dialog_reprogramar_cita()
    if st.session_state.get("_ap_fin_item"):
        _dialog_ajustar_montos()
    if st.session_state.get("_ap_cancel_item"):
        _dialog_cancelar_cita()
    if st.session_state.get("_ap_receipts_item"):
        _dialog_recibos_cita()

    st.markdown("##### Reporte")
    # st.tabs ejecuta cada pestaña en cada rerun; el radio solo dibuja una rama.
    rep_sec = st.radio(
        "Sección",
        ["Finanzas — citas", "Encuestas — satisfacción"],
        horizontal=True,
        key="rep_subsection",
    )
    if rep_sec.startswith("Finanzas"):
        _render_reporte_financiero_citas_body(items, svc_values, status_values)
    else:
        st.markdown("##### Resumen por pregunta")
        with st.spinner("Cargando estadísticas de encuesta…"):
            _render_survey_question_stats_report()


def render_citas_tab() -> None:
    """Calendario, agendar y diálogo de citas del día; datos financieros y tabla en **Reporte**."""
    _init_appt_tab_session_state()
    _consume_cal_appt_query_param()
    _inject_citas_shared_styles()
    _sync_appointments_from_api()

    _render_appointments_fetch_error_toast()
    _render_appointment_action_feedback()

    items = list(st.session_state.get("_ap_list") or [])
    svc_values = list(st.session_state.get("_ap_svc_values") or [])
    status_values = ["Agendada", "Reprogramada", "Finalizada", "Cancelada"]

    st.markdown("##### Gestión citas — calendario")

    with st.expander(
        "Leyenda de colores — citas",
        expanded=True,
    ):
        _render_citas_color_legend()

    filt_head_l, filt_head_r = st.columns([4.2, 1.35])
    with filt_head_l:
        st.markdown("##### Filtros del calendario")
    with filt_head_r:
        _cex_disabled = _panel_is_technician_role()
        if st.button(
            "Cita express",
            type="primary",
            use_container_width=True,
            disabled=_cex_disabled,
            key="btn_cita_express_cal_header",
            help=(
                "Piercing: agenda con abono completo, alta de cliente, encuesta de piercing y firma del contrato."
                if not _cex_disabled
                else "Los tatuadores y perforadores no pueden iniciar este flujo desde aquí."
            ),
        ):
            open_contract_express_piercing()
    cf1, cf2, cf3 = st.columns([1.3, 1.0, 1.0])
    with cf1:
        st.text_input("Filtrar nombre", key="_ap_cal_f_name", placeholder="Nombre cliente")
    with cf2:
        st.selectbox("Servicio", options=["Todos", *svc_values], key="_ap_cal_f_service")
    with cf3:
        st.selectbox("Estado", options=["Todos", *status_values], key="_ap_cal_f_status")

    _render_professional_calendar_filter()

    vista_compact = "Mes — compacta"
    vista_team = "Mes — por equipo"
    vista_week = "Semana (rejilla)"

    fid = int(st.session_state.get("_ap_filter_artist_id") or 0)
    may_all = _may_see_all_appointments()
    can_team = may_all and fid == 0
    vista_options: list[str] = [vista_compact] + ([vista_team] if can_team else []) + [vista_week]

    period_key = "_ap_cal_period"
    layout_key = "_ap_cal_layout"
    if period_key not in st.session_state:
        st.session_state[period_key] = "Mes"

    if "_ap_cal_vista" not in st.session_state:
        pk = str(st.session_state.get(period_key) or "Mes")
        lk = str(st.session_state.get(layout_key) or "Compacta")
        if pk == "Semana (rejilla)":
            st.session_state["_ap_cal_vista"] = vista_week
        elif lk == "Por equipo" and can_team:
            st.session_state["_ap_cal_vista"] = vista_team
        else:
            st.session_state["_ap_cal_vista"] = vista_compact

    vista = str(st.session_state.get("_ap_cal_vista") or vista_compact)
    if vista == vista_team and not can_team:
        vista = vista_compact
        st.session_state["_ap_cal_vista"] = vista_compact
    if vista not in vista_options:
        vista = vista_compact
        st.session_state["_ap_cal_vista"] = vista_compact

    prev_vista = st.session_state.get("__ap_cal_vista_prev")
    if vista == vista_week and prev_vista != vista_week:
        _clear_calendar_dialog_focus()
        _sync_week_monday_for_agenda_context()

    st.radio(
        "Vista calendario",
        vista_options,
        horizontal=True,
        key="_ap_cal_vista",
        help=(
            "**Mes compacta**: filas del día sin agrupar. "
            "**Mes por equipo**: agrupa por profesional cuando el filtro permite ver a todos; "
            "si cambias el filtro, vuelves a compacta. "
            "**Semana**: columnas Lun–Dom con franjas de **30 min** (*Outlook / Teams*). "
            "Al cambiar desde el mes a la vista semana, la primera semana se alinea con el mes abierto."
        ),
    )

    vista = str(st.session_state.get("_ap_cal_vista") or vista_compact)
    if vista == vista_team and not can_team:
        vista = vista_compact
        st.session_state["_ap_cal_vista"] = vista_compact
    if vista not in vista_options:
        vista = vista_compact
        st.session_state["_ap_cal_vista"] = vista_compact

    st.session_state["__ap_cal_vista_prev"] = vista

    team_layout = False
    if vista == vista_week:
        period_sel = "Semana (rejilla)"
        st.markdown(
            "<div style=\"opacity:.58;font-size:0.74rem;line-height:1.35;"
            'padding-top:0.08rem;margin-bottom:0.35rem">Semana compacta · <strong>Lun–Dom</strong> · franjas 30 min'
            "</div>",
            unsafe_allow_html=True,
        )
    elif vista == vista_team:
        period_sel = "Mes"
        team_layout = True
        st.session_state[layout_key] = "Por equipo"
        st.session_state[period_key] = "Mes"
    else:
        period_sel = "Mes"
        st.session_state[layout_key] = "Compacta"
        st.session_state[period_key] = "Mes"

    cal_filtered = _apply_appointment_filters(
        items,
        use_date_range=False,
        name_key="_ap_cal_f_name",
        service_key="_ap_cal_f_service",
        status_key="_ap_cal_f_status",
    )
    hc_raw = st.session_state.get("_ap_hist_counts")
    hist_counts: dict[str, int] = dict(hc_raw) if isinstance(hc_raw, dict) else {}
    if not hist_counts and items:
        hist_counts = _appointment_counts_by_client(items)
    by_day = _appointments_by_day_sorted(cal_filtered)

    _invoke_citas_tab_dialogs(by_day, hist_counts)

    if period_sel == "Semana (rejilla)":
        _render_week_schedule_grid(by_day, hist_counts)
    else:
        _render_main_calendar(by_day, hist_counts, team_layout=team_layout)
